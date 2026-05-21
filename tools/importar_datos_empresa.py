#!/usr/bin/env python
"""
Importa datos históricos desde los Excel de la empresa al ERP Universal.

Archivos fuente (Datos Generales de la empresa/):
  - CLIENTES - ESTADO DE CUENTAS 2023-2024 (1).xlsx
  - CLIENTES - ESTADO DE CUENTAS 2025 (Autoguardado) (1).xlsx
  - COSTOS MP - INSUMOS - PROD TERMINADOS (1).xlsx

Destino: %APPDATA%/ERP Universal/erp.db  (o --db PATH para otro archivo)

Uso:
  python tools/importar_datos_empresa.py
  python tools/importar_datos_empresa.py --db C:/ruta/erp.db
  python tools/importar_datos_empresa.py --gen-seed   # genera seed_empresa.py también
"""
import sys
import os
import re
import uuid
import sqlite3
import argparse
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("Error: falta openpyxl. Instalar con:  pip install openpyxl")

# ─── Rutas ────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
DATOS_DIR = ROOT / "Datos Generales de la empresa"

EXCEL_CUENTAS_2324 = DATOS_DIR / "CLIENTES - ESTADO DE CUENTAS 2023-2024 (1).xlsx"
EXCEL_CUENTAS_2526 = DATOS_DIR / "CLIENTES - ESTADO DE CUENTAS 2025 (Autoguardado) (1).xlsx"
EXCEL_COSTOS       = DATOS_DIR / "COSTOS MP - INSUMOS - PROD TERMINADOS (1).xlsx"

FACTURA_RE = re.compile(r'\d{3}-\d{3}-\d+')


# ─── Helpers ──────────────────────────────────────────────────────────────────

def gen_id() -> str:
    return str(uuid.uuid4())


def limpia(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Normaliza espacios múltiples
    return re.sub(r'\s+', ' ', s)


def es_factura(val) -> bool:
    return bool(val and FACTURA_RE.search(limpia(val)))


def to_date_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = limpia(val)
    if s and re.match(r'\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return None


def to_decimal(val, default="0") -> str:
    if val is None:
        return default
    try:
        return str(round(float(val), 2))
    except (ValueError, TypeError):
        return default


# ─── Lectura de Excel ─────────────────────────────────────────────────────────

def leer_clientes_y_comprobantes():
    """
    Devuelve:
      clientes: dict  nombre_norm → nombre_display
      comprobantes: list[dict]
    """
    clientes: dict[str, str] = {}
    comprobantes: list[dict] = []

    def _norm(nombre: str) -> str:
        return nombre.upper().strip()

    def _procesar_sheet(ws, tiene_pendiente: bool):
        for row in ws.iter_rows(min_row=2, values_only=True):
            col0 = row[0]
            col1 = row[1] if len(row) > 1 else None
            col2 = row[2] if len(row) > 2 else None
            col3 = row[3] if len(row) > 3 else None
            col7 = row[7] if len(row) > 7 else None

            if not es_factura(col0):
                continue
            if not col2:   # sin monto → ignorar
                continue

            nombre_cliente = limpia(col1) if col1 else ""
            if not nombre_cliente:
                continue

            fecha_str = to_date_str(col3)
            if not fecha_str:
                continue

            norm = _norm(nombre_cliente)
            if norm not in clientes:
                clientes[norm] = nombre_cliente

            monto = to_decimal(col2)
            pendiente = to_decimal(col7) if tiene_pendiente and col7 else "0"

            comprobantes.append({
                "numero":    limpia(col0),
                "cliente":   norm,
                "monto":     monto,
                "fecha":     fecha_str,
                "pendiente": pendiente,
            })

    # 2023-2024
    if EXCEL_CUENTAS_2324.exists():
        wb = openpyxl.load_workbook(str(EXCEL_CUENTAS_2324), data_only=True)
        if "Hoja1" in wb.sheetnames:
            _procesar_sheet(wb["Hoja1"], tiene_pendiente=False)

    # 2024-2025 y 2025-2026
    if EXCEL_CUENTAS_2526.exists():
        wb = openpyxl.load_workbook(str(EXCEL_CUENTAS_2526), data_only=True)
        for sheet in ["VENTAS 2024-2025", "VENTAS 2025-2026"]:
            if sheet in wb.sheetnames:
                _procesar_sheet(wb[sheet], tiene_pendiente=True)

    return clientes, comprobantes


def leer_inventario():
    """
    Devuelve:
      proveedores: dict  nombre_norm → nombre_display
      mp:    list[dict]  — Materias Primas
      ins:   list[dict]  — Insumos y Envases
      pt:    list[dict]  — Productos Terminados
    """
    proveedores: dict[str, str] = {}
    mp: list[dict] = []
    ins: list[dict] = []
    pt: list[dict] = []

    if not EXCEL_COSTOS.exists():
        return proveedores, mp, ins, pt

    wb = openpyxl.load_workbook(str(EXCEL_COSTOS), data_only=True)

    # ─ Materias Primas ────────────────────────────────────────────────────────
    if "MP STOCK VALORIZADO" in wb.sheetnames:
        ws = wb["MP STOCK VALORIZADO"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            codigo = row[0] if row[0] else None
            if not codigo or not isinstance(codigo, (int, float)):
                continue
            nombre = limpia(row[1]) if row[1] else ""
            if not nombre:
                continue
            cantidad = to_decimal(row[2], "0")
            proveedor = limpia(row[5]) if len(row) > 5 and row[5] else ""
            costo = to_decimal(row[6]) if len(row) > 6 else "0"

            if proveedor:
                norm = proveedor.upper()
                # Normalizar M CASSAB / MCASSAB como mismo proveedor
                if "CASSAB" in norm:
                    proveedor = "M CASSAB"
                    norm = "M CASSAB"
                if norm not in proveedores:
                    proveedores[norm] = proveedor

            mp.append({
                "codigo":    str(int(codigo)),
                "descripcion": nombre,
                "cantidad":  cantidad,
                "costo":     costo,
                "proveedor": proveedor.upper() if proveedor else "",
            })

    # ─ Insumos y Envases ─────────────────────────────────────────────────────
    if "INSUMOS" in wb.sheetnames:
        ws = wb["INSUMOS"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            codigo = row[0] if row[0] else None
            if not codigo or not isinstance(codigo, (int, float)):
                continue
            nombre = limpia(row[1]) if row[1] else ""
            if not nombre:
                continue
            cantidad = to_decimal(row[2], "0")
            proveedor = limpia(row[3]) if len(row) > 3 and row[3] else ""
            costo = to_decimal(row[4]) if len(row) > 4 else "0"

            if proveedor:
                norm = proveedor.upper()
                if norm not in proveedores:
                    proveedores[norm] = proveedor

            ins.append({
                "codigo":    str(int(codigo)),
                "descripcion": nombre,
                "cantidad":  cantidad,
                "costo":     costo,
                "proveedor": proveedor.upper() if proveedor else "",
            })

    # ─ Productos Terminados ───────────────────────────────────────────────────
    if "PROD TERMINADOS" in wb.sheetnames:
        ws = wb["PROD TERMINADOS"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            codigo = row[1] if len(row) > 1 else None
            if not codigo or not isinstance(codigo, (int, float)):
                continue
            nombre = limpia(row[2]) if len(row) > 2 and row[2] else ""
            if not nombre:
                continue
            costo = to_decimal(row[3]) if len(row) > 3 else "0"

            pt.append({
                "codigo":    str(int(codigo)),
                "descripcion": nombre,
                "cantidad":  "0",
                "costo":     costo,
                "proveedor": "",
            })

    return proveedores, mp, ins, pt


# ─── Inserción en SQLite ──────────────────────────────────────────────────────

def insertar_todos(db_path: str, verbose: bool = True) -> dict:
    """Inserta todos los datos en la DB. Devuelve resumen con conteos."""
    # isolation_level=None = autocommit: cada execute() es atómico, sin estado inconsistente
    conn = sqlite3.connect(db_path, isolation_level=None)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.create_function("uuid_generate_v4", 0, lambda: str(uuid.uuid4()))

    # ─ empresa_id
    row = conn.execute(
        "SELECT id FROM empresas ORDER BY fecha_creacion ASC LIMIT 1"
    ).fetchone()
    if not row:
        conn.close()
        raise RuntimeError("No existe ninguna empresa en la DB. Arrancá el ERP primero.")
    empresa_id = row[0]

    # ─ tipos de comprobante
    tipo_venta = conn.execute(
        "SELECT id FROM tipos_comprobante WHERE nombre LIKE '%Venta%' AND empresa_id=?",
        (empresa_id,)
    ).fetchone()
    if not tipo_venta:
        tipo_venta_id = gen_id()
        conn.execute(
            "INSERT INTO tipos_comprobante (id, empresa_id, nombre) VALUES (?,?,?)",
            (tipo_venta_id, empresa_id, "Factura de Venta")
        )
    else:
        tipo_venta_id = tipo_venta[0]

    # ─ Leer Excel
    if verbose:
        print("Leyendo Excel...")
    clientes_raw, comprobantes_raw = leer_clientes_y_comprobantes()
    proveedores_raw, mp_items, ins_items, pt_items = leer_inventario()

    stats = {
        "clientes": 0, "proveedores": 0,
        "categorias": 0, "inventario": 0,
        "comprobantes": 0,
    }

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

    # ─ Clientes
    cliente_ids: dict[str, str] = {}  # norm → uuid
    for norm, display in clientes_raw.items():
        cid = gen_id()
        try:
            conn.execute(
                """INSERT INTO clientes (id, empresa_id, nombre, activo, fecha_creacion)
                   VALUES (?, ?, ?, 1, ?)""",
                (cid, empresa_id, display, now)
            )
            cliente_ids[norm] = cid
            stats["clientes"] += 1
        except sqlite3.IntegrityError:
            existing = conn.execute(
                "SELECT id FROM clientes WHERE empresa_id=? AND nombre=?",
                (empresa_id, display)
            ).fetchone()
            if existing:
                cliente_ids[norm] = existing[0]

    # ─ Proveedores
    proveedor_ids: dict[str, str] = {}  # norm → uuid
    for norm, display in proveedores_raw.items():
        pid = gen_id()
        try:
            conn.execute(
                """INSERT INTO proveedores (id, empresa_id, nombre, activo, fecha_creacion)
                   VALUES (?, ?, ?, 1, ?)""",
                (pid, empresa_id, display, now)
            )
            proveedor_ids[norm] = pid
            stats["proveedores"] += 1
        except sqlite3.IntegrityError:
            existing = conn.execute(
                "SELECT id FROM proveedores WHERE empresa_id=? AND nombre=?",
                (empresa_id, display)
            ).fetchone()
            if existing:
                proveedor_ids[norm] = existing[0]

    # ─ Categorías de inventario
    categorias = [
        ("Materias Primas",    "Materias primas para producción"),
        ("Insumos y Envases",  "Insumos, envases y materiales de packaging"),
        ("Productos Terminados", "Stock de productos terminados listos para venta"),
    ]
    cat_ids: dict[str, str] = {}
    for nombre_cat, desc_cat in categorias:
        cid = gen_id()
        try:
            conn.execute(
                """INSERT INTO categorias_inventario (id, empresa_id, nombre, descripcion)
                   VALUES (?, ?, ?, ?)""",
                (cid, empresa_id, nombre_cat, desc_cat)
            )
            cat_ids[nombre_cat] = cid
            stats["categorias"] += 1
        except sqlite3.IntegrityError:
            existing = conn.execute(
                "SELECT id FROM categorias_inventario WHERE empresa_id=? AND nombre=?",
                (empresa_id, nombre_cat)
            ).fetchone()
            if existing:
                cat_ids[nombre_cat] = existing[0]

    # ─ Inventario
    def insertar_items(items: list[dict], cat_nombre: str, unidad: str):
        cat_id = cat_ids.get(cat_nombre)
        for item in items:
            iid = gen_id()
            try:
                conn.execute(
                    """INSERT INTO inventario
                       (id, empresa_id, categoria_id, codigo, descripcion,
                        unidad_medida, cantidad_actual, costo_unitario, punto_reorden, activo, fecha_creacion)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, 1, ?)""",
                    (iid, empresa_id, cat_id, item["codigo"],
                     item["descripcion"], unidad,
                     item["cantidad"], item["costo"], now)
                )
                stats["inventario"] += 1
            except sqlite3.IntegrityError:
                pass  # código duplicado → ignorar

    insertar_items(mp_items,  "Materias Primas",     "kg")
    insertar_items(ins_items, "Insumos y Envases",   "unidad")
    insertar_items(pt_items,  "Productos Terminados","unidad")

    # ─ Comprobantes de venta
    for comp in comprobantes_raw:
        norm_cliente = comp["cliente"]
        cliente_id = cliente_ids.get(norm_cliente)
        if not cliente_id:
            continue

        monto = comp["monto"]
        pendiente = comp["pendiente"]
        # IVA 10% incluido si no se conoce el desglose
        monto_f = float(monto)
        iva = round(monto_f * 10 / 110, 2)
        subtotal = round(monto_f - iva, 2)

        cid = gen_id()
        try:
            conn.execute(
                """INSERT INTO comprobantes
                   (id, empresa_id, tipo_id, numero_comprobante, fecha_emision,
                    cliente_id, monto_subtotal, monto_iva, monto_total, saldo_pendiente,
                    metodo_carga, estado_validacion, condicion, fecha_creacion, fecha_modificacion)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual', 'confirmado', 'credito', ?, ?)""",
                (cid, empresa_id, tipo_venta_id,
                 comp["numero"], comp["fecha"],
                 cliente_id,
                 str(subtotal), str(iva), monto, pendiente,
                 now, now)
            )
            stats["comprobantes"] += 1
            # Detalle mínimo (una línea resumen)
            conn.execute(
                """INSERT INTO detalle_comprobantes
                   (id, empresa_id, comprobante_id, descripcion,
                    cantidad, precio_unitario, porcentaje_iva, subtotal, iva_monto)
                   VALUES (?, ?, ?, ?, 1, ?, 10, ?, ?)""",
                (gen_id(), empresa_id, cid,
                 "Venta (importado desde Excel historico)",
                 monto, str(subtotal), str(iva))
            )
        except (sqlite3.IntegrityError, sqlite3.OperationalError):
            pass  # comprobante ya existe o FK issue → ignorar

    conn.close()
    return stats


# ─── Generador de seed_empresa.py ─────────────────────────────────────────────

def generar_seed_py(db_path: str) -> None:
    """
    Lee los datos importados de la DB y genera backend/core/seed_empresa.py
    para que nuevas instalaciones del .exe también tengan los datos.
    """
    conn = sqlite3.connect(db_path)
    empresa_id_row = conn.execute(
        "SELECT id FROM empresas ORDER BY fecha_creacion ASC LIMIT 1"
    ).fetchone()
    if not empresa_id_row:
        conn.close()
        print("WARN: no hay empresa en la DB para generar seed.")
        return

    clientes = conn.execute(
        "SELECT nombre FROM clientes WHERE empresa_id=? ORDER BY nombre",
        (empresa_id_row[0],)
    ).fetchall()
    proveedores = conn.execute(
        "SELECT nombre FROM proveedores WHERE empresa_id=? ORDER BY nombre",
        (empresa_id_row[0],)
    ).fetchall()
    categorias = conn.execute(
        "SELECT nombre, descripcion FROM categorias_inventario WHERE empresa_id=? ORDER BY nombre",
        (empresa_id_row[0],)
    ).fetchall()
    inventario = conn.execute(
        """SELECT i.codigo, i.descripcion, i.unidad_medida, i.cantidad_actual,
                  i.costo_unitario, c.nombre as cat
           FROM inventario i
           LEFT JOIN categorias_inventario c ON c.id = i.categoria_id
           WHERE i.empresa_id=?
           ORDER BY c.nombre, i.codigo""",
        (empresa_id_row[0],)
    ).fetchall()
    comprobantes = conn.execute(
        """SELECT c.numero_comprobante, cl.nombre, c.monto_total, c.saldo_pendiente,
                  c.fecha_emision
           FROM comprobantes c
           JOIN clientes cl ON cl.id = c.cliente_id
           WHERE c.empresa_id=?
           ORDER BY c.fecha_emision""",
        (empresa_id_row[0],)
    ).fetchall()
    conn.close()

    seed_path = ROOT / "backend" / "core" / "seed_empresa.py"

    lines = [
        '"""',
        'Datos semilla pre-cargados de la empresa.',
        'Generado automáticamente por tools/importar_datos_empresa.py',
        'Se aplica una sola vez al primer arranque del ERP (sin clientes previos).',
        '"""',
        'from __future__ import annotations',
        'import uuid as _uuid',
        'import sqlite3',
        'from datetime import datetime as _dt, timezone as _tz',
        '',
        '',
        'CLIENTES = [',
    ]
    for (nombre,) in clientes:
        lines.append(f'    {repr(nombre)},')
    lines += [
        ']',
        '',
        'PROVEEDORES = [',
    ]
    for (nombre,) in proveedores:
        lines.append(f'    {repr(nombre)},')
    lines += [
        ']',
        '',
        'CATEGORIAS = [',
    ]
    for nombre, desc in categorias:
        lines.append(f'    ({repr(nombre)}, {repr(desc)}),')
    lines += [
        ']',
        '',
        '# (codigo, descripcion, unidad_medida, cantidad_actual, costo_unitario, categoria)',
        'INVENTARIO = [',
    ]
    for row in inventario:
        lines.append(f'    {tuple(row)!r},')
    lines += [
        ']',
        '',
        '# (numero_comprobante, nombre_cliente, monto_total, saldo_pendiente, fecha_emision)',
        'COMPROBANTES = [',
    ]
    for row in comprobantes:
        lines.append(f'    {tuple(row)!r},')
    lines += [
        ']',
        '',
        '',
        'def sembrar_sqlite(conn: sqlite3.Connection, empresa_id: str) -> None:',
        '    """Inserta datos semilla si la tabla clientes está vacía."""',
        '    existing = conn.execute(',
        '        "SELECT count(*) FROM clientes WHERE empresa_id=?", (empresa_id,)',
        '    ).fetchone()[0]',
        '    if existing > 0:',
        '        return',
        '',
        '    now = _dt.now(_tz.utc).strftime("%Y-%m-%d %H:%M:%S")',
        '',
        '    # Clientes',
        '    cliente_ids: dict[str, str] = {}',
        '    for nombre in CLIENTES:',
        '        cid = str(_uuid.uuid4())',
        '        conn.execute(',
        '            "INSERT OR IGNORE INTO clientes (id, empresa_id, nombre, activo, fecha_creacion)"',
        '            " VALUES (?,?,?,1,?)",',
        '            (cid, empresa_id, nombre, now)',
        '        )',
        '        cliente_ids[nombre] = cid',
        '',
        '    # Proveedores',
        '    for nombre in PROVEEDORES:',
        '        conn.execute(',
        '            "INSERT OR IGNORE INTO proveedores (id, empresa_id, nombre, activo, fecha_creacion)"',
        '            " VALUES (?,?,?,1,?)",',
        '            (str(_uuid.uuid4()), empresa_id, nombre, now)',
        '        )',
        '',
        '    # Categorías inventario',
        '    cat_ids: dict[str, str] = {}',
        '    for nombre, desc in CATEGORIAS:',
        '        cid = str(_uuid.uuid4())',
        '        conn.execute(',
        '            "INSERT OR IGNORE INTO categorias_inventario (id, empresa_id, nombre, descripcion)"',
        '            " VALUES (?,?,?,?)",',
        '            (cid, empresa_id, nombre, desc)',
        '        )',
        '        cat_ids[nombre] = cid',
        '',
        '    # Inventario',
        '    for codigo, desc, unidad, cantidad, costo, cat in INVENTARIO:',
        '        cat_id = cat_ids.get(cat)',
        '        conn.execute(',
        '            "INSERT OR IGNORE INTO inventario"',
        '            " (id, empresa_id, categoria_id, codigo, descripcion, unidad_medida,"',
        '            "  cantidad_actual, costo_unitario, punto_reorden, activo, fecha_creacion)"',
        '            " VALUES (?,?,?,?,?,?,?,?,0,1,?)",',
        '            (str(_uuid.uuid4()), empresa_id, cat_id, codigo, desc, unidad, cantidad, costo, now)',
        '        )',
        '',
        '    # Tipos comprobante',
        '    tipo_venta = conn.execute(',
        '        "SELECT id FROM tipos_comprobante WHERE nombre LIKE \'%Venta%\' AND empresa_id=?",',
        '        (empresa_id,)',
        '    ).fetchone()',
        '    if not tipo_venta:',
        '        return',
        '    tipo_venta_id = tipo_venta[0]',
        '',
        '    # Comprobantes',
        '    for num, nombre_cl, monto, pendiente, fecha in COMPROBANTES:',
        '        cl_id = cliente_ids.get(nombre_cl)',
        '        if not cl_id:',
        '            cl_id = conn.execute(',
        '                "SELECT id FROM clientes WHERE empresa_id=? AND nombre=?",',
        '                (empresa_id, nombre_cl)',
        '            ).fetchone()',
        '            if cl_id:',
        '                cl_id = cl_id[0]',
        '            else:',
        '                continue',
        '        monto_f = float(monto)',
        '        iva = round(monto_f * 10 / 110, 2)',
        '        subtotal = round(monto_f - iva, 2)',
        '        cid = str(_uuid.uuid4())',
        '        conn.execute(',
        '            "INSERT OR IGNORE INTO comprobantes"',
        '            " (id, empresa_id, tipo_id, numero_comprobante, fecha_emision,"',
        '            "  cliente_id, monto_subtotal, monto_iva, monto_total, saldo_pendiente,"',
        '            "  metodo_carga, estado_validacion, condicion, fecha_creacion, fecha_modificacion)"',
        '            " VALUES (?,?,?,?,?,?,?,?,?,?,\'manual\',\'confirmado\',\'credito\',?,?)",',
        '            (cid, empresa_id, tipo_venta_id, num, fecha, cl_id,',
        '             str(subtotal), str(iva), str(monto_f), str(pendiente), now, now)',
        '        )',
        '',
        '    conn.commit()',
    ]

    seed_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"Seed generado: {seed_path}")


# ─── Punto de entrada ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Importar datos empresa al ERP Universal")
    parser.add_argument("--db", help="Ruta al archivo erp.db (default: %%APPDATA%%/ERP Universal/erp.db)")
    parser.add_argument("--gen-seed", action="store_true",
                        help="Generar backend/core/seed_empresa.py tras importar")
    args = parser.parse_args()

    if args.db:
        db_path = args.db
    else:
        appdata = os.environ.get("APPDATA") or str(Path.home())
        db_path = str(Path(appdata) / "ERP Universal" / "erp.db")

    if not Path(db_path).exists():
        sys.exit(f"No se encontró la DB en: {db_path}\n"
                 f"Arrancá el ERP al menos una vez para crear la base de datos.")

    print(f"DB: {db_path}")
    print("Importando datos...")

    stats = insertar_todos(db_path, verbose=True)

    print("\n=== Resumen de importacion ===========================")
    print(f"  Clientes insertados:     {stats['clientes']}")
    print(f"  Proveedores insertados:  {stats['proveedores']}")
    print(f"  Categorias inventario:   {stats['categorias']}")
    print(f"  Items inventario:        {stats['inventario']}")
    print(f"  Comprobantes (facturas): {stats['comprobantes']}")
    print("======================================================")
    print("Importacion completada.")

    if args.gen_seed:
        print("\nGenerando seed_empresa.py...")
        generar_seed_py(db_path)

    return stats


if __name__ == "__main__":
    main()
