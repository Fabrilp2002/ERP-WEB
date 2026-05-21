"""
Migra datos productivos desde los Excel de la empresa a Supabase Postgres.

Lee:
  - CLIENTES - ESTADO DE CUENTAS 2025 (Autoguardado) (1).xlsx (sheets VENTAS 2024-2025, VENTAS 2025-2026)
  - CLIENTES - ESTADO DE CUENTAS 2023-2024 (1).xlsx (sheets Hoja1, Hoja2)
  - COSTOS MP - INSUMOS - PROD TERMINADOS (1).xlsx (sheets MP, INSUMOS, PROD TERMINADOS)

Carga en este orden (con dedup, idempotente):
  1. Clientes (lookup por nombre)
  2. Proveedores (lookup por nombre)
  3. Categorías de inventario (Materia Prima, Insumo, Producto Terminado)
  4. Inventario (lookup por código)
  5. Comprobantes de venta (con cliente_id y fecha; SIN items detallados — la data
     fuente solo trae monto total)

Uso:
    python scripts/migrar_excels.py [--dry-run] [--excel-dir "C:/ruta/a/excels"]

Requiere:
    backend/.env con DATABASE_URL configurada (apuntar al pooler de Supabase).
"""
from __future__ import annotations
import argparse
import asyncio
import os
import re
import sys
from pathlib import Path
from datetime import date, datetime
from typing import Iterable

import asyncpg
import openpyxl

# ─── Config ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_DIR = Path(r"C:\Users\gfcar\Desktop\IA\Empresa 1\Datos Generales de la empresa")
EXCEL_DIR = Path(os.environ.get("ERP_EXCEL_DIR", DEFAULT_EXCEL_DIR))

XLS_VENTAS_25 = EXCEL_DIR / "CLIENTES - ESTADO DE CUENTAS 2025 (Autoguardado) (1).xlsx"
XLS_VENTAS_23 = EXCEL_DIR / "CLIENTES - ESTADO DE CUENTAS 2023-2024 (1).xlsx"
XLS_INV       = EXCEL_DIR / "COSTOS MP - INSUMOS - PROD TERMINADOS (1).xlsx"


def configurar_excel_dir(excel_dir: str | Path | None) -> None:
    """Permite usar los mismos scripts fuera de la PC original del PM."""
    global EXCEL_DIR, XLS_VENTAS_25, XLS_VENTAS_23, XLS_INV
    if excel_dir:
        EXCEL_DIR = Path(excel_dir).expanduser().resolve()
    XLS_VENTAS_25 = EXCEL_DIR / "CLIENTES - ESTADO DE CUENTAS 2025 (Autoguardado) (1).xlsx"
    XLS_VENTAS_23 = EXCEL_DIR / "CLIENTES - ESTADO DE CUENTAS 2023-2024 (1).xlsx"
    XLS_INV = EXCEL_DIR / "COSTOS MP - INSUMOS - PROD TERMINADOS (1).xlsx"


def validar_excel_dir() -> None:
    faltantes = [p for p in (XLS_VENTAS_25, XLS_VENTAS_23, XLS_INV) if not p.exists()]
    if faltantes:
        nombres = "\n  - ".join(str(p) for p in faltantes)
        sys.exit(
            "ERROR: no encontre los Excel fuente. Usa --excel-dir o ERP_EXCEL_DIR.\n"
            f"Faltantes:\n  - {nombres}"
        )


# ─── Utilidades ──────────────────────────────────────────────────────────────

def cargar_database_url() -> str:
    """Lee DATABASE_URL del .env del backend o de la env."""
    if "DATABASE_URL" in os.environ:
        return os.environ["DATABASE_URL"]
    env_file = ROOT / "backend" / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            if line.startswith("DATABASE_URL="):
                v = line.split("=", 1)[1].strip().strip('"').strip("'")
                # asyncpg quiere postgresql:// (sin +asyncpg)
                return v.replace("postgresql+asyncpg://", "postgresql://", 1)
    sys.exit("ERROR: DATABASE_URL no encontrada en env ni backend/.env")


def es_nombre_cliente_real(s: str | None) -> bool:
    """Filtra strings que NO son nombres reales de clientes (productos, comentarios, etc)."""
    if not s:
        return False
    s = s.strip()
    if len(s) < 3 or len(s) > 100:
        return False
    upper = s.upper()
    blacklist_starts = (
        "ESPLENDIDA", "TROCHE", "SALDO", "DESCRIPCIÓN", "DESCRIPCION",
        "TOTAL", "FACTURA", "RECIBO", "NOTA", "PRODUCTO", "CODIGO",
        "CÓDIGO", "MATERIA",
        "ASUNCIÓN", "ASUNCION", "CONSIGNACIÓN", "CONSIGNACION",
        "OBSERVAC", "ME FACTURARON", "NDC", "NO TENGO", "VARIOS",
    )
    if any(upper.startswith(b) for b in blacklist_starts):
        return False
    blacklist_contains = (
        "OBSERVACIÓN", "OBSERVACION", "ME FACTURARON", "SE ANULÓ", "SE ANULO",
        "NO TENGO", "POR DESCUENTO", "DESCRIPCIÓN DEL PRODUCTO",
    )
    if any(b in upper for b in blacklist_contains):
        return False
    if upper in ("HOJA1", "HOJA2"):
        return False
    if re.match(r"^\d", s):  # empieza con número
        return False
    # Que tenga al menos una letra
    if not re.search(r"[a-zA-Z]", s):
        return False
    return True


def es_proveedor_real(s: str | None) -> bool:
    if not s:
        return False
    s = s.strip()
    if len(s) < 3 or len(s) > 100:
        return False
    upper = s.upper()
    blacklist = ("PROVEEDOR", "FILTRADO", "TOTAL", "VARIOS")
    if any(upper.startswith(b) for b in blacklist):
        return False
    return True


def normalizar_nombre(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip())


def agregar_item_inventario(items: list[dict], item: dict) -> None:
    """Consolida filas repetidas del Excel cuando comparten codigo de inventario."""
    existente = next((x for x in items if x["codigo"] == item["codigo"]), None)
    if not existente:
        items.append(item)
        return

    existente["cantidad"] = float(existente.get("cantidad") or 0) + float(item.get("cantidad") or 0)
    if not existente.get("precio_compra") and item.get("precio_compra"):
        existente["precio_compra"] = item["precio_compra"]
    if not existente.get("precio_venta") and item.get("precio_venta"):
        existente["precio_venta"] = item["precio_venta"]
    if not existente.get("proveedor") and item.get("proveedor"):
        existente["proveedor"] = item["proveedor"]


# ─── Extractores por archivo ─────────────────────────────────────────────────

def extraer_facturas_ventas() -> tuple[set[str], list[dict]]:
    """
    Recorre los 3 sheets de ventas (2023-2024 + 2024-2025 + 2025-2026)
    y devuelve (clientes_unicos, lista_facturas).
    """
    clientes: set[str] = set()
    facturas: list[dict] = []

    sources = [
        (XLS_VENTAS_25, "VENTAS 2024-2025"),
        (XLS_VENTAS_25, "VENTAS 2025-2026"),
        (XLS_VENTAS_23, "Hoja1"),
        (XLS_VENTAS_23, "Hoja2"),
    ]

    for archivo, sheet_name in sources:
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb[sheet_name]
        current_cliente: str | None = None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                continue  # header
            row = row + (None,) * 15

            # Detectar layout por columnas: usar las primeras 8 columnas
            # y para Hoja2 (que tiene offset de 1 col vacía) compensamos
            offset = 1 if sheet_name == "Hoja2" else 0
            fact     = row[0 + offset]
            cliente  = row[1 + offset]
            monto    = row[2 + offset]
            fecha    = row[3 + offset]
            recibo   = row[4 + offset]
            unidades = row[5 + offset] if "2023-2024" in archivo.name else None

            # 1) Fila de "encabezado de cliente": col A es nombre, otras vacías
            if isinstance(fact, str) and not cliente and not monto:
                if es_nombre_cliente_real(fact):
                    current_cliente = normalizar_nombre(fact)
                    clientes.add(current_cliente)
                continue

            # 2) Fila con datos
            cliente_real = None
            if cliente and es_nombre_cliente_real(str(cliente)):
                cliente_real = normalizar_nombre(str(cliente))
                clientes.add(cliente_real)
                current_cliente = cliente_real
            elif current_cliente:
                cliente_real = current_cliente

            if not cliente_real or not monto:
                continue
            try:
                monto_f = float(monto)
            except (TypeError, ValueError):
                continue
            if monto_f <= 0:
                continue

            # Limpiar número de factura
            num_fact = None
            if fact and isinstance(fact, str):
                f = fact.strip()
                if f.startswith("00") or "001-" in f:
                    num_fact = re.sub(r"\s+", "-", f)
                    num_fact = re.sub(r"-+", "-", num_fact)

            # Fecha
            fecha_iso = None
            if isinstance(fecha, (date, datetime)):
                fecha_iso = fecha.strftime("%Y-%m-%d") if hasattr(fecha, "strftime") else str(fecha)[:10]
            elif isinstance(fecha, str):
                fecha_iso = fecha[:10]

            facturas.append({
                "cliente": cliente_real,
                "numero": num_fact,
                "monto": monto_f,
                "fecha": fecha_iso,
                "fuente": f"{archivo.name}::{sheet_name}",
            })

        wb.close()

    return clientes, facturas


def extraer_inventario() -> tuple[set[str], list[dict]]:
    """Devuelve (proveedores_unicos, lista_items_inventario)."""
    proveedores: set[str] = set()
    items: list[dict] = []

    wb = openpyxl.load_workbook(XLS_INV, read_only=True, data_only=True)

    # MP
    ws = wb["MP STOCK VALORIZADO"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 4:
            continue
        cod, nom, cant, lote, vence, prov, costo_kg, *_ = row + (None,) * 9
        if not cod or not nom:
            continue
        try:
            cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
        except (TypeError, ValueError):
            continue
        nombre = normalizar_nombre(str(nom))
        prov_str = normalizar_nombre(str(prov)) if prov and es_proveedor_real(str(prov)) else None
        if prov_str:
            proveedores.add(prov_str)
        agregar_item_inventario(items, {
            "codigo": cod_str,
            "nombre": nombre,
            "categoria": "Materia Prima",
            "cantidad": float(cant) if isinstance(cant, (int, float)) else 0.0,
            "precio_compra": float(costo_kg) if isinstance(costo_kg, (int, float)) else None,
            "precio_venta": None,
            "proveedor": prov_str,
        })

    # Insumos
    ws = wb["INSUMOS"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 3:
            continue
        cod, nom, cant, prov, precio, *_ = row + (None,) * 5
        if not cod or not nom:
            continue
        try:
            cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
        except (TypeError, ValueError):
            continue
        nombre = normalizar_nombre(str(nom))
        prov_str = normalizar_nombre(str(prov)) if prov and es_proveedor_real(str(prov)) else None
        if prov_str:
            proveedores.add(prov_str)
        agregar_item_inventario(items, {
            "codigo": cod_str,
            "nombre": nombre,
            "categoria": "Insumo",
            "cantidad": float(cant) if isinstance(cant, (int, float)) else 0.0,
            "precio_compra": float(precio) if isinstance(precio, (int, float)) else None,
            "precio_venta": None,
            "proveedor": prov_str,
        })

    # Productos terminados
    ws = wb["PROD TERMINADOS"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 5:
            continue
        a, cod, nom, costo, p_may, p_min, *_ = row + (None,) * 7
        if not cod or not nom:
            continue
        try:
            cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
        except (TypeError, ValueError):
            continue
        nombre = normalizar_nombre(str(nom))
        agregar_item_inventario(items, {
            "codigo": cod_str,
            "nombre": nombre,
            "categoria": "Producto Terminado",
            "cantidad": 0.0,
            "precio_compra": float(costo) if isinstance(costo, (int, float)) else None,
            "precio_venta": float(p_min or p_may) if isinstance(p_min or p_may, (int, float)) else None,
            "proveedor": None,
        })

    wb.close()
    return proveedores, items


# ─── Cargas a BD ─────────────────────────────────────────────────────────────

async def get_empresa_id(conn) -> str:
    eid = await conn.fetchval(
        "SELECT id FROM empresas ORDER BY fecha_creacion ASC LIMIT 1"
    )
    if not eid:
        sys.exit("ERROR: no hay empresa en la BD. Correr scripts/bootstrap_admin.py primero.")
    return str(eid)


async def get_tipo_factura_venta(conn, empresa_id: str) -> str:
    tid = await conn.fetchval(
        "SELECT id FROM tipos_comprobante WHERE empresa_id=$1 AND nombre ILIKE '%venta%' LIMIT 1",
        empresa_id,
    )
    if not tid:
        sys.exit("ERROR: no encontré tipo 'Factura de Venta' en tipos_comprobante")
    return str(tid)


async def upsert_clientes(conn, empresa_id: str, nombres: Iterable[str], dry: bool):
    creados = actualizados = 0
    for nombre in sorted(nombres):
        existente = await conn.fetchval(
            "SELECT id FROM clientes WHERE empresa_id=$1 AND nombre=$2",
            empresa_id, nombre,
        )
        if existente:
            if not dry:
                await conn.execute(
                    """
                    UPDATE inventario
                    SET codigo=$2, descripcion=$3, categoria_id=$4, cantidad_actual=$5,
                        costo_unitario=$6, unidad_medida=$7, activo=TRUE
                    WHERE id=$8 AND empresa_id=$1
                    """,
                    empresa_id, item["codigo"], item["nombre"], cat_id,
                    item["cantidad"], item["precio_compra"] or 0,
                    "kg" if item["categoria"] == "Materia Prima" else "unidad",
                    existente,
                )
            actualizados += 1
            continue
        if dry:
            creados += 1
            continue
        await conn.execute(
            """INSERT INTO clientes (empresa_id, nombre, ruc) VALUES ($1, $2, NULL)""",
            empresa_id, nombre,
        )
        creados += 1
    return creados, actualizados


async def upsert_proveedores(conn, empresa_id: str, nombres: Iterable[str], dry: bool):
    creados = actualizados = 0
    for nombre in sorted(nombres):
        existente = await conn.fetchval(
            "SELECT id FROM proveedores WHERE empresa_id=$1 AND nombre=$2",
            empresa_id, nombre,
        )
        if existente:
            actualizados += 1
            continue
        if dry:
            creados += 1
            continue
        await conn.execute(
            """INSERT INTO proveedores (empresa_id, nombre, ruc) VALUES ($1, $2, NULL)""",
            empresa_id, nombre,
        )
        creados += 1
    return creados, actualizados


async def upsert_categorias_inventario(conn, empresa_id: str, dry: bool) -> dict[str, str]:
    """Devuelve dict {nombre_categoria: id}."""
    nombres = ["Materia Prima", "Insumo", "Producto Terminado"]
    result: dict[str, str] = {}
    for nombre in nombres:
        cid = await conn.fetchval(
            "SELECT id FROM categorias_inventario WHERE empresa_id=$1 AND nombre=$2",
            empresa_id, nombre,
        )
        if not cid and not dry:
            cid = await conn.fetchval(
                "INSERT INTO categorias_inventario (empresa_id, nombre) VALUES ($1, $2) RETURNING id",
                empresa_id, nombre,
            )
        if cid:
            result[nombre] = str(cid)
    return result


async def upsert_inventario(conn, empresa_id: str, items: list[dict], cat_ids: dict[str, str], dry: bool):
    creados = actualizados = saltados = 0
    for item in items:
        cat_id = cat_ids.get(item["categoria"])
        if not cat_id:
            saltados += 1
            continue
        # Lookup por código
        existente = await conn.fetchval(
            "SELECT id FROM inventario WHERE empresa_id=$1 AND codigo=$2",
            empresa_id, item["codigo"],
        )
        if existente:
            actualizados += 1
            continue
        if dry:
            creados += 1
            continue
        try:
            await conn.execute(
                """
                INSERT INTO inventario
                  (empresa_id, codigo, descripcion, categoria_id, cantidad_actual, costo_unitario, unidad_medida, activo)
                VALUES ($1, $2, $3, $4, $5, $6, $7, TRUE)
                """,
                empresa_id, item["codigo"], item["nombre"], cat_id,
                item["cantidad"], item["precio_compra"] or 0,
                "kg" if item["categoria"] == "Materia Prima" else "unidad",
            )
            creados += 1
        except Exception as e:
            print(f"  SKIP {item['codigo']} ({item['nombre'][:30]}): {str(e)[:100]}")
            saltados += 1
    return creados, actualizados, saltados


async def upsert_facturas(conn, empresa_id: str, tipo_id: str, facturas: list[dict],
                          productos_por_cliente: dict, dry: bool):
    """
    Inserta cada factura + UN detalle (no hay desglose de items en los Excel
    fuente, solo el monto total). El detalle lleva una descripción rica con
    el contexto del Excel y los productos típicos de ese cliente cuando se
    conocen (de la matriz VENTAS 2025-2026).
    Cálculo IVA: asume 10% incluido en el total → subtotal=total/1.1, iva=total/11.
    """
    creadas = actualizadas = saltadas = 0
    for f in facturas:
        if not f["fecha"]:
            saltadas += 1
            continue
        # Validar que fecha sea ISO real (YYYY-MM-DD)
        try:
            fecha_dt = date.fromisoformat(f["fecha"][:10])
        except (ValueError, TypeError):
            saltadas += 1
            continue
        f["fecha_dt"] = fecha_dt  # cachear el date parseado
        cliente_id = await conn.fetchval(
            "SELECT id FROM clientes WHERE empresa_id=$1 AND nombre=$2",
            empresa_id, f["cliente"],
        )
        if not cliente_id:
            saltadas += 1
            continue

        # Dedup
        if f["numero"]:
            existente = await conn.fetchval(
                "SELECT id FROM comprobantes WHERE empresa_id=$1 AND numero_comprobante=$2",
                empresa_id, f["numero"],
            )
        else:
            existente = await conn.fetchval(
                """SELECT id FROM comprobantes
                   WHERE empresa_id=$1 AND cliente_id=$2 AND fecha_emision=$3 AND monto_total=$4""",
                empresa_id, cliente_id, f["fecha_dt"], f["monto"],
            )
        if existente:
            actualizadas += 1
            continue
        if dry:
            creadas += 1
            continue

        num = f["numero"] or f"SIN-NUM-{f['fecha']}-{int(f['monto'])}"
        monto = f["monto"]
        subtotal = round(monto / 1.10, 2)        # 10% IVA incluido
        iva_monto = round(monto - subtotal, 2)

        # Descripción del ítem: enriquecemos con productos típicos del cliente
        desc_principal = f"Venta facturada — Factura {num}" if f['numero'] else f"Venta facturada al cliente"
        productos_cli = productos_por_cliente.get(f["cliente"], [])
        if productos_cli:
            top3 = ", ".join(f"{nombre} ({u} ud.)" for nombre, u in productos_cli[:3])
            desc_principal += f". Productos típicos: {top3}"

        notas = (
            f"Importado desde Excel: {f['fuente']}. "
            f"Monto incluye IVA 10% (subtotal: ₲{int(subtotal):,}, IVA: ₲{int(iva_monto):,})."
        )

        try:
            comp_id = await conn.fetchval(
                """
                INSERT INTO comprobantes
                  (empresa_id, tipo_id, numero_comprobante, fecha_emision, cliente_id,
                   monto_subtotal, monto_iva, monto_total, saldo_pendiente,
                   metodo_carga, estado_validacion, condicion, notas)
                VALUES ($1, $2, $3, $4, $5,
                        $6, $7, $8, $8,
                        'manual', 'confirmado', 'credito', $9)
                RETURNING id
                """,
                empresa_id, tipo_id, num, f["fecha_dt"], cliente_id,
                subtotal, iva_monto, monto, notas,
            )
            # Detalle de la factura
            await conn.execute(
                """
                INSERT INTO detalle_comprobantes
                  (empresa_id, comprobante_id, descripcion, cantidad,
                   precio_unitario, porcentaje_iva, subtotal, iva_monto)
                VALUES ($1, $2, $3, 1, $4, 10, $5, $6)
                """,
                empresa_id, comp_id, desc_principal[:500],
                subtotal, subtotal, iva_monto,
            )
            creadas += 1
        except Exception as e:
            saltadas += 1
            if saltadas <= 5:
                print(f"  SKIP factura {num}: {str(e)[:120]}")

    return creadas, actualizadas, saltadas


def extraer_productos_por_cliente() -> dict[str, list[tuple[str, int]]]:
    """
    Lee la hoja 'VENTAS 2025-2026' del archivo de inventario que es una matriz
    productos × clientes con cantidades. Devuelve {cliente_normalizado: [(producto, cantidad), ...]}
    ordenado por cantidad descendente.
    """
    wb = openpyxl.load_workbook(XLS_INV, read_only=True, data_only=True)
    ws = wb["VENTAS 2025-2026"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return {}

    # Buscar la fila de encabezados de clientes (R1 según vimos antes)
    # Las columnas con clientes: 4=ALIMENTOS, 8=CAFSA, 12=COSCOM, 15=FARMA, 16=MARTIN, 17=BELEN, 18=CADENA
    # Vamos a tomar el header de R1 + R3 (sublocación) para crear la columna→cliente
    header_r1 = rows[0] if rows else ()
    cliente_por_col: dict[int, str] = {}
    last_cliente = None
    for col_idx, val in enumerate(header_r1):
        if val and isinstance(val, str) and len(val.strip()) > 2:
            last_cliente = normalizar_nombre(val)
        if last_cliente:
            cliente_por_col[col_idx] = last_cliente

    # Iterar productos (desde R4 en adelante; col 1=código, col 2=descripción)
    result: dict[str, list[tuple[str, int]]] = {}
    for row in rows[3:]:
        cod = row[1] if len(row) > 1 else None
        nom = row[2] if len(row) > 2 else None
        if not nom or not isinstance(nom, str):
            continue
        producto = normalizar_nombre(nom)
        for col_idx, qty in enumerate(row):
            if col_idx <= 3 or not isinstance(qty, (int, float)) or qty <= 0:
                continue
            cli = cliente_por_col.get(col_idx)
            if not cli:
                continue
            result.setdefault(cli, []).append((producto, int(qty)))

    # Ordenar por cantidad descendente
    for cli in result:
        result[cli].sort(key=lambda x: -x[1])
    wb.close()
    return result


# ─── Main ────────────────────────────────────────────────────────────────────

def extraer_productos_por_cliente() -> dict[str, list[tuple[str, int]]]:
    """
    Lee la matriz productos x clientes y excluye columnas de totales/costos.
    Reemplaza la version inicial para evitar que columnas como VENTAS/COSTO
    queden asignadas por herencia al ultimo cliente visible.
    """
    wb = openpyxl.load_workbook(XLS_INV, read_only=True, data_only=True)
    ws = wb["VENTAS 2025-2026"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        wb.close()
        return {}

    header_r1 = rows[0] if rows else ()
    header_r3 = rows[2] if len(rows) > 2 else ()
    cliente_por_col: dict[int, str] = {}
    last_cliente = None
    for col_idx, val in enumerate(header_r1):
        if isinstance(val, str) and len(val.strip()) > 2:
            last_cliente = normalizar_nombre(val)
        if last_cliente and col_idx >= 4:
            cliente_por_col[col_idx] = last_cliente

    blacklist_r3 = ("VENTAS", "COSTO", "TOTAL", "PRODUCC")
    cols_validas: set[int] = set()
    for col_idx in cliente_por_col:
        if col_idx >= len(header_r3):
            continue
        sucursal = header_r3[col_idx]
        if not isinstance(sucursal, str) or not sucursal.strip():
            continue
        if any(token in sucursal.upper() for token in blacklist_r3):
            continue
        cols_validas.add(col_idx)

    result: dict[str, list[tuple[str, int]]] = {}
    for row in rows[3:]:
        nom = row[2] if len(row) > 2 else None
        if not isinstance(nom, str) or not nom.strip():
            continue
        if "VENTAS" in nom.upper() or "COSTO" in nom.upper():
            continue
        producto = normalizar_nombre(nom)
        if len(producto) < 5:
            continue
        for col_idx, qty in enumerate(row):
            if col_idx not in cols_validas:
                continue
            if not isinstance(qty, (int, float)) or qty <= 0:
                continue
            if qty > 10000:
                continue
            cli = cliente_por_col.get(col_idx)
            if cli:
                result.setdefault(cli, []).append((producto, int(qty)))

    for cli in list(result.keys()):
        acumulado: dict[str, int] = {}
        for prod, qty in result[cli]:
            acumulado[prod] = acumulado.get(prod, 0) + qty
        result[cli] = sorted(acumulado.items(), key=lambda x: -x[1])

    wb.close()
    return result


async def main(dry: bool, excel_dir: str | None = None):
    print(f"{'═' * 60}")
    print(f"  MIGRACIÓN DE EXCEL A SUPABASE  {'(DRY RUN)' if dry else '(REAL)'}")
    print(f"{'═' * 60}")

    configurar_excel_dir(excel_dir)
    validar_excel_dir()
    print(f"\n→ Excel dir: {EXCEL_DIR}")

    db_url = cargar_database_url()
    print(f"\n→ Conectando a: {db_url.split('@')[-1].split('/')[0] if '@' in db_url else '?'}")
    conn = await asyncpg.connect(db_url, ssl="require")

    try:
        empresa_id = await get_empresa_id(conn)
        tipo_id = await get_tipo_factura_venta(conn, empresa_id)
        print(f"  empresa_id = {empresa_id}")
        print(f"  tipo_id (factura venta) = {tipo_id}\n")

        # 1. Extraer datos
        print("→ Leyendo Excels...")
        clientes_set, facturas = extraer_facturas_ventas()
        proveedores_set, items_inv = extraer_inventario()
        productos_x_cliente = extraer_productos_por_cliente()
        print(f"  clientes únicos:           {len(clientes_set)}")
        print(f"  proveedores únicos:        {len(proveedores_set)}")
        print(f"  items inventario:          {len(items_inv)}")
        print(f"  facturas a cargar:         {len(facturas)}")
        print(f"  clientes con matriz prod.: {len(productos_x_cliente)}")

        # 2. Cargas
        print("\n→ Cargando clientes...")
        c, a = await upsert_clientes(conn, empresa_id, clientes_set, dry)
        print(f"  creados={c}  ya existían={a}")

        print("\n→ Cargando proveedores...")
        c, a = await upsert_proveedores(conn, empresa_id, proveedores_set, dry)
        print(f"  creados={c}  ya existían={a}")

        print("\n→ Cargando categorías de inventario...")
        cat_ids = await upsert_categorias_inventario(conn, empresa_id, dry)
        print(f"  categorías: {list(cat_ids.keys())}")

        print("\n→ Cargando inventario...")
        c, a, s = await upsert_inventario(conn, empresa_id, items_inv, cat_ids, dry)
        print(f"  creados={c}  ya existían={a}  saltados={s}")

        print("\n→ Cargando comprobantes/facturas (con detalle de items)...")
        c, a, s = await upsert_facturas(conn, empresa_id, tipo_id, facturas, productos_x_cliente, dry)
        print(f"  creadas={c}  ya existían={a}  saltadas={s}")

        # 3. Resumen final
        print(f"\n{'═' * 60}")
        print(f"  Conteo final en BD")
        print(f"{'═' * 60}")
        for tabla in ("clientes", "proveedores", "inventario", "comprobantes", "categorias_inventario"):
            n = await conn.fetchval(
                f"SELECT COUNT(*) FROM {tabla} WHERE empresa_id=$1", empresa_id,
            )
            print(f"  {tabla:25s}: {n}")
        if dry:
            print("\n[DRY RUN] Nada se guardó. Correr sin --dry-run para aplicar cambios.")
        else:
            print("\n[OK] Migración completada.")
    finally:
        await conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="No persiste, solo cuenta")
    parser.add_argument("--excel-dir", help="Carpeta que contiene los 3 Excel fuente")
    args = parser.parse_args()
    asyncio.run(main(dry=args.dry_run, excel_dir=args.excel_dir))
