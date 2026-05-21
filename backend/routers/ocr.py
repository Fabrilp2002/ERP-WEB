"""
Router de carga de facturas — procesamiento automático con Gemini Vision.
"""
import base64
import re
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal
from typing import Literal, Optional
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from ..core.database import get_db
from ..core.security import get_current_user, require_escritura
from ..services.ocr import extraer_datos_factura
from ..core import key_store

router = APIRouter(prefix="/ocr", tags=["Facturas"])

MIME_PERMITIDOS = {
    "image/jpeg", "image/png", "image/webp", "image/gif",
    "application/pdf",
}
MAX_SIZE_MB = 20

# MIME types aceptados para la importación masiva desde Excel
MIME_EXCEL = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls (legacy — openpyxl sólo lee xlsx)
    "application/octet-stream",  # algunos browsers lo mandan así
}


@router.post("/preview", summary="Vista previa de PDF o imagen")
async def preview_archivo(
    archivo: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Convierte cada página de un PDF a imagen PNG para mostrar en pantalla."""
    contenido = await archivo.read()
    mime = archivo.content_type or ""
    imagenes = []

    if mime == "application/pdf":
        try:
            import fitz
            doc = fitz.open(stream=contenido, filetype="pdf")
            for i, page in enumerate(doc):
                if i >= 4:
                    break
                mat = fitz.Matrix(1.5, 1.5)
                pix = page.get_pixmap(matrix=mat)
                png_bytes = pix.tobytes("png")
                imagenes.append({
                    "pagina": i + 1,
                    "data": f"data:image/png;base64,{base64.b64encode(png_bytes).decode()}",
                    "ancho": pix.width,
                    "alto": pix.height,
                })
            doc.close()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error procesando PDF: {str(e)}")
    elif mime.startswith("image/"):
        imagenes.append({
            "pagina": 1,
            "data": f"data:{mime};base64,{base64.b64encode(contenido).decode()}",
        })
    else:
        raise HTTPException(status_code=422, detail="Formato no soportado")

    return {"imagenes": imagenes, "total_paginas": len(imagenes)}


@router.get("/status", summary="Estado de configuración")
async def ocr_status(current_user: dict = Depends(get_current_user)):
    """Devuelve si Gemini está configurado."""
    return {"gemini_configurado": key_store.is_configured()}


@router.post("/extraer", summary="Extraer datos de factura (acepta múltiples archivos)")
async def extraer_factura(
    archivos: list[UploadFile] = File(..., description="Una o más imágenes/PDFs de la MISMA factura"),
    current_user: dict = Depends(require_escritura),
):
    """
    Sube una o varias imágenes/PDFs de la MISMA factura (ej. frente + dorso,
    o una foto por página) y devuelve un único JSON combinado.
    NO crea el comprobante — el operador debe revisar y confirmar.
    """
    if not archivos:
        raise HTTPException(status_code=422, detail="No se recibió ningún archivo")

    archivos_preparados: list[tuple[bytes, str]] = []
    total_mb = 0.0
    for arch in archivos:
        if arch.content_type not in MIME_PERMITIDOS:
            raise HTTPException(
                status_code=422,
                detail=f"Tipo no soportado: {arch.content_type}. Usar: JPEG, PNG, WebP, PDF",
            )
        contenido = await arch.read()
        size_mb = len(contenido) / (1024 * 1024)
        if size_mb > MAX_SIZE_MB:
            raise HTTPException(
                status_code=413,
                detail=f"Archivo '{arch.filename}' muy grande ({size_mb:.1f}MB). Máx: {MAX_SIZE_MB}MB",
            )
        total_mb += size_mb
        archivos_preparados.append((contenido, arch.content_type or "image/jpeg"))

    if total_mb > MAX_SIZE_MB * 2:
        raise HTTPException(
            status_code=413,
            detail=f"Total combinado muy grande ({total_mb:.1f}MB). Máx combinado: {MAX_SIZE_MB * 2}MB",
        )

    resultado = await extraer_datos_factura(archivos_preparados)

    if "error" in resultado:
        raise HTTPException(status_code=502, detail=resultado["error"])

    return resultado


@router.post("/procesar", summary="Extraer + crear comprobante pendiente")
async def procesar_factura(
    archivo: UploadFile = File(...),
    tipo_id: UUID = Query(..., description="UUID del tipo de comprobante"),
    cliente_id: UUID = Query(None, description="UUID del cliente (si es venta)"),
    proveedor_id: UUID = Query(None, description="UUID del proveedor (si es compra)"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_escritura),
):
    """Extrae datos y crea comprobante en estado pendiente_revision."""
    empresa_id = current_user["empresa_id"]
    usuario_id = current_user["sub"]

    if archivo.content_type not in MIME_PERMITIDOS:
        raise HTTPException(status_code=422, detail=f"Tipo no soportado: {archivo.content_type}")
    if not cliente_id and not proveedor_id:
        raise HTTPException(status_code=422, detail="Debe indicar cliente_id o proveedor_id")
    if cliente_id and proveedor_id:
        raise HTTPException(status_code=422, detail="No puede tener cliente Y proveedor")

    contenido = await archivo.read()
    size_mb = len(contenido) / (1024 * 1024)
    if size_mb > MAX_SIZE_MB:
        raise HTTPException(status_code=413, detail=f"Archivo muy grande ({size_mb:.1f}MB)")

    datos = await extraer_datos_factura(
        archivo_bytes=contenido,
        mime_type=archivo.content_type or "image/jpeg",
    )

    if "error" in datos:
        raise HTTPException(status_code=502, detail=datos["error"])

    monto_total = Decimal(str(datos.get("monto_total", 0)))
    monto_subtotal = Decimal(str(datos.get("monto_subtotal", 0)))
    monto_iva_5 = Decimal(str(datos.get("monto_iva_5", 0)))
    monto_iva_10 = Decimal(str(datos.get("monto_iva_10", 0)))
    monto_iva = monto_iva_5 + monto_iva_10

    metodo = "ocr_imagen" if "image" in (archivo.content_type or "") else "ocr_pdf"

    # Convertir fecha string → date (asyncpg no acepta strings para DATE)
    fecha_str = datos.get("fecha_emision")
    fecha_obj = None
    if fecha_str:
        try:
            fecha_obj = datetime.strptime(str(fecha_str).strip(), "%Y-%m-%d").date()
        except ValueError:
            fecha_obj = None
    if not fecha_obj:
        fecha_obj = date.today()

    result = await db.execute(
        text("""
            INSERT INTO comprobantes (
                empresa_id, tipo_id, numero_comprobante, fecha_emision,
                cliente_id, proveedor_id,
                monto_subtotal, monto_iva, monto_total, saldo_pendiente,
                metodo_carga, estado_validacion, usuario_carga_id, notas
            ) VALUES (
                :empresa_id, :tipo_id, :numero, :fecha,
                :cliente_id, :proveedor_id,
                :subtotal, :iva, :total, :total,
                :metodo, 'confirmado', :usuario_id,
                :notas
            )
            RETURNING id, empresa_id, numero_comprobante, fecha_emision,
                      monto_total, saldo_pendiente, metodo_carga,
                      estado_validacion, fecha_creacion
        """),
        {
            "empresa_id": empresa_id,
            "tipo_id": str(tipo_id),
            "numero": datos.get("numero_comprobante", "PENDIENTE"),
            "fecha": fecha_obj,
            "cliente_id": str(cliente_id) if cliente_id else None,
            "proveedor_id": str(proveedor_id) if proveedor_id else None,
            "subtotal": monto_subtotal,
            "iva": monto_iva,
            "total": monto_total,
            "metodo": metodo,
            "usuario_id": usuario_id,
            "notas": f"Procesado automáticamente — Confianza: {datos.get('confianza', '?')}",
        },
    )
    comprobante = dict(result.mappings().first())

    for item in datos.get("items", []):
        cantidad = Decimal(str(item.get("cantidad", 1)))
        precio = Decimal(str(item.get("precio_unitario", 0)))
        iva_pct = Decimal(str(item.get("porcentaje_iva", 10)))
        subtotal = (cantidad * precio).quantize(Decimal("0.01"))
        if iva_pct == Decimal("10"):
            iva_monto = (subtotal / Decimal("11")).quantize(Decimal("0.01"))
        elif iva_pct == Decimal("5"):
            iva_monto = (subtotal / Decimal("21")).quantize(Decimal("0.01"))
        else:
            iva_monto = Decimal("0.00")

        await db.execute(
            text("""
                INSERT INTO detalle_comprobantes (
                    empresa_id, comprobante_id, descripcion,
                    cantidad, precio_unitario, porcentaje_iva,
                    subtotal, iva_monto
                ) VALUES (
                    :empresa_id, :comprobante_id, :desc,
                    :cantidad, :precio, :iva_pct,
                    :subtotal, :iva_monto
                )
            """),
            {
                "empresa_id": empresa_id,
                "comprobante_id": str(comprobante["id"]),
                "desc": item.get("descripcion", "Ítem sin descripción"),
                "cantidad": cantidad,
                "precio": precio,
                "iva_pct": iva_pct,
                "subtotal": subtotal,
                "iva_monto": iva_monto,
            },
        )

    return {
        "comprobante": comprobante,
        "datos_extraidos": datos,
        "mensaje": "Comprobante creado como pendiente_revision.",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINT UNIFICADO: Confirmar datos extraídos y guardar todo en una sola llamada
# (auto-crea cliente/proveedor si no existe, resuelve tipo_id automáticamente)
# ═══════════════════════════════════════════════════════════════════════════════

class ItemConfirmar(BaseModel):
    descripcion: str = ""
    cantidad: float = 1
    precio_unitario: float = 0
    porcentaje_iva: int = 10
    codigo: Optional[str] = None


class ConfirmarFacturaIn(BaseModel):
    tipo: Literal["venta", "compra"]
    numero_comprobante: str
    fecha_emision: Optional[str] = None  # YYYY-MM-DD
    ruc_emisor: Optional[str] = None
    razon_social_emisor: Optional[str] = None
    ruc_cliente: Optional[str] = None
    razon_social_cliente: Optional[str] = None
    condicion: Literal["contado", "credito"] = "contado"
    medio_pago_contado: Optional[Literal["efectivo", "transferencia", "cheque", "tarjeta", "otro"]] = None
    fecha_vencimiento: Optional[str] = None  # YYYY-MM-DD (solo credito)
    plazo_dias: Optional[int] = None  # alternativa: dias desde fecha_emision
    ubicacion_fisica: Optional[str] = None  # lugar donde esta archivada en papel
    monto_subtotal: float = 0
    monto_iva_5: float = 0
    monto_iva_10: float = 0
    monto_total: float = 0
    items: list[ItemConfirmar] = Field(default_factory=list)
    confianza: Optional[float] = None
    motor_usado: Optional[str] = None


def _limpiar_ruc(ruc: Optional[str]) -> Optional[str]:
    """Normaliza un RUC: saca espacios, reemplaza guiones raros por el normal."""
    if not ruc:
        return None
    r = ruc.strip().replace(" ", "").replace("–", "-").replace("—", "-")
    return r or None


async def _obtener_tipo_id(db: AsyncSession, empresa_id: str, nombre: str) -> str:
    """Devuelve el UUID del tipo_comprobante. Lo crea si no existe."""
    res = await db.execute(
        text("""
            SELECT id FROM tipos_comprobante
            WHERE empresa_id = :empresa_id AND nombre = :nombre
            LIMIT 1
        """),
        {"empresa_id": empresa_id, "nombre": nombre},
    )
    row = res.mappings().first()
    if row:
        return str(row["id"])
    # Crear si no existe (tolerante en empresas nuevas sin seed)
    res = await db.execute(
        text("""
            INSERT INTO tipos_comprobante (empresa_id, nombre)
            VALUES (:empresa_id, :nombre)
            RETURNING id
        """),
        {"empresa_id": empresa_id, "nombre": nombre},
    )
    return str(res.mappings().first()["id"])


async def _resolver_cliente(
    db: AsyncSession, empresa_id: str, nombre: Optional[str], ruc: Optional[str], usuario_id: str
) -> Optional[str]:
    """Busca cliente por RUC o nombre. Si no existe, lo crea. Devuelve UUID o None."""
    ruc = _limpiar_ruc(ruc)
    nombre = (nombre or "").strip() or None
    if not ruc and not nombre:
        return None

    # Buscar por RUC primero (más confiable), luego por nombre
    if ruc:
        res = await db.execute(
            text("SELECT id FROM clientes WHERE empresa_id = :e AND ruc = :r LIMIT 1"),
            {"e": empresa_id, "r": ruc},
        )
        row = res.mappings().first()
        if row:
            return str(row["id"])
    if nombre:
        res = await db.execute(
            text("SELECT id FROM clientes WHERE empresa_id = :e AND LOWER(nombre) = LOWER(:n) LIMIT 1"),
            {"e": empresa_id, "n": nombre},
        )
        row = res.mappings().first()
        if row:
            return str(row["id"])

    # No existe → crearlo
    res = await db.execute(
        text("""
            INSERT INTO clientes (empresa_id, nombre, ruc, notas)
            VALUES (:e, :n, :r, 'Creado automáticamente desde carga de factura')
            RETURNING id
        """),
        {"e": empresa_id, "n": nombre or f"Cliente {ruc}", "r": ruc},
    )
    return str(res.mappings().first()["id"])


async def _resolver_proveedor(
    db: AsyncSession, empresa_id: str, nombre: Optional[str], ruc: Optional[str], usuario_id: str
) -> Optional[str]:
    """Busca proveedor por RUC o nombre. Si no existe, lo crea. Devuelve UUID o None."""
    ruc = _limpiar_ruc(ruc)
    nombre = (nombre or "").strip() or None
    if not ruc and not nombre:
        return None

    if ruc:
        res = await db.execute(
            text("SELECT id FROM proveedores WHERE empresa_id = :e AND ruc = :r LIMIT 1"),
            {"e": empresa_id, "r": ruc},
        )
        row = res.mappings().first()
        if row:
            return str(row["id"])
    if nombre:
        res = await db.execute(
            text("SELECT id FROM proveedores WHERE empresa_id = :e AND LOWER(nombre) = LOWER(:n) LIMIT 1"),
            {"e": empresa_id, "n": nombre},
        )
        row = res.mappings().first()
        if row:
            return str(row["id"])

    res = await db.execute(
        text("""
            INSERT INTO proveedores (empresa_id, nombre, ruc, notas)
            VALUES (:e, :n, :r, 'Creado automáticamente desde carga de factura')
            RETURNING id
        """),
        {"e": empresa_id, "n": nombre or f"Proveedor {ruc}", "r": ruc},
    )
    return str(res.mappings().first()["id"])


async def _resolver_articulo(
    db: AsyncSession,
    empresa_id: str,
    codigo: Optional[str],
    descripcion: Optional[str],
) -> tuple[Optional[str], Optional[str], str]:
    """
    Cruza código ↔ descripción contra la tabla `inventario` de la empresa.

    Si el OCR leyó sólo el código (porque la descripción estaba ilegible), buscamos
    el artículo por código y tomamos su descripción oficial del catálogo.
    Si el OCR leyó sólo la descripción (código ilegible), buscamos por descripción
    (coincidencia exacta primero, luego ILIKE) y tomamos el código oficial.
    Si ambos vienen, priorizamos la coincidencia por código y sólo reemplazamos
    descripción si la del catálogo difiere significativamente (informativo).

    Devuelve (inventario_id, codigo_final, descripcion_final).
    """
    cod = (codigo or "").strip() or None
    desc = (descripcion or "").strip() or None

    # 1) Match por código exacto
    if cod:
        res = await db.execute(
            text("""
                SELECT id, codigo, descripcion
                FROM inventario
                WHERE empresa_id = :e AND codigo = :c AND activo = TRUE
                LIMIT 1
            """),
            {"e": empresa_id, "c": cod},
        )
        row = res.mappings().first()
        if row:
            # Si la descripción venía vacía o ilegible, usamos la del catálogo
            return (
                str(row["id"]),
                row["codigo"],
                desc or row["descripcion"] or "Ítem",
            )

    # 2) Match por descripción exacta (case-insensitive)
    if desc:
        res = await db.execute(
            text("""
                SELECT id, codigo, descripcion
                FROM inventario
                WHERE empresa_id = :e AND LOWER(descripcion) = LOWER(:d) AND activo = TRUE
                LIMIT 1
            """),
            {"e": empresa_id, "d": desc},
        )
        row = res.mappings().first()
        if row:
            return (str(row["id"]), cod or row["codigo"], row["descripcion"])

        # 3) Match por descripción aproximada (prefijo de 12+ chars)
        if len(desc) >= 6:
            res = await db.execute(
                text("""
                    SELECT id, codigo, descripcion
                    FROM inventario
                    WHERE empresa_id = :e AND activo = TRUE
                      AND descripcion ILIKE :p
                    LIMIT 2
                """),
                {"e": empresa_id, "p": f"%{desc[:20]}%"},
            )
            matches = list(res.mappings())
            if len(matches) == 1:
                row = matches[0]
                return (str(row["id"]), cod or row["codigo"], row["descripcion"])

    # 4) No match → devolvemos lo que vino (al menos descripción no vacía)
    return (None, cod, desc or "Ítem")


@router.get("/articulo-lookup", summary="Buscar artículo en inventario por código o descripción")
async def articulo_lookup(
    codigo: Optional[str] = None,
    descripcion: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Cruce en vivo código ↔ descripción contra inventario.
    Usado por la UI de carga de facturas para autocompletar el campo vacío
    cuando el operador escribe solo uno de los dos.
    """
    empresa_id = current_user["empresa_id"]
    if not codigo and not descripcion:
        return {"encontrado": False}
    inv_id, cod, desc = await _resolver_articulo(db, empresa_id, codigo, descripcion)
    return {
        "encontrado": inv_id is not None,
        "inventario_id": inv_id,
        "codigo": cod,
        "descripcion": desc,
    }


@router.post("/confirmar", summary="Confirmar y guardar factura extraída (todo en uno)")
async def confirmar_factura(
    datos: ConfirmarFacturaIn,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_escritura),
):
    """
    Guarda un comprobante ya revisado por el usuario.
    - Auto-crea cliente/proveedor si el RUC no existe en la empresa
    - Resuelve tipo_id desde el nombre ('Factura de Venta' / 'Factura de Compra')
    - Crea el comprobante en estado pendiente_revision
    - Inserta items del detalle
    """
    empresa_id = current_user["empresa_id"]
    usuario_id = current_user["sub"]

    if not datos.numero_comprobante.strip():
        raise HTTPException(status_code=422, detail="Falta el número de comprobante")
    if not datos.fecha_emision:
        raise HTTPException(status_code=422, detail="Falta la fecha de emisión")

    # Convertir fecha de string "YYYY-MM-DD" a date (asyncpg lo exige para DATE)
    try:
        fecha_obj: date = datetime.strptime(datos.fecha_emision.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail=f"Fecha inválida: '{datos.fecha_emision}'. Formato esperado: YYYY-MM-DD",
        )

    # 1) Resolver tipo_id
    tipo_nombre = "Factura de Venta" if datos.tipo == "venta" else "Factura de Compra"
    tipo_id = await _obtener_tipo_id(db, empresa_id, tipo_nombre)

    # 2) Resolver cliente o proveedor según tipo
    cliente_id: Optional[str] = None
    proveedor_id: Optional[str] = None

    if datos.tipo == "venta":
        # En una VENTA, la contraparte es el CLIENTE (campo "Nombre o Razón Social")
        cliente_id = await _resolver_cliente(
            db, empresa_id, datos.razon_social_cliente, datos.ruc_cliente, usuario_id
        )
        if not cliente_id:
            raise HTTPException(
                status_code=422,
                detail="No se pudo identificar al cliente. Completá al menos el RUC o la razón social.",
            )
    else:  # compra
        # En una COMPRA, la contraparte es el PROVEEDOR (quien emite la factura)
        proveedor_id = await _resolver_proveedor(
            db, empresa_id, datos.razon_social_emisor, datos.ruc_emisor, usuario_id
        )
        if not proveedor_id:
            raise HTTPException(
                status_code=422,
                detail="No se pudo identificar al proveedor. Completá al menos el RUC o la razón social.",
            )

    # 3) Calcular montos (confiamos en los del form; si el usuario los editó, usamos esos)
    monto_total = Decimal(str(datos.monto_total or 0))
    monto_subtotal = Decimal(str(datos.monto_subtotal or 0))
    monto_iva = Decimal(str((datos.monto_iva_5 or 0) + (datos.monto_iva_10 or 0)))

    # 4) Calcular fecha_vencimiento si es credito
    fecha_venc_obj = None
    if datos.condicion == "credito":
        if datos.fecha_vencimiento:
            try:
                fecha_venc_obj = date.fromisoformat(datos.fecha_vencimiento)
            except Exception:
                fecha_venc_obj = None
        if fecha_venc_obj is None and datos.plazo_dias and fecha_obj:
            from datetime import timedelta
            fecha_venc_obj = fecha_obj + timedelta(days=int(datos.plazo_dias))

    # 4b) Insertar comprobante
    notas = f"Carga automática — Condición: {datos.condicion} — Confianza: {int((datos.confianza or 0) * 100)}%"
    try:
        result = await db.execute(
            text("""
                INSERT INTO comprobantes (
                    empresa_id, tipo_id, numero_comprobante, fecha_emision, fecha_vencimiento,
                    cliente_id, proveedor_id,
                    monto_subtotal, monto_iva, monto_total, saldo_pendiente,
                    metodo_carga, condicion, medio_pago_contado, ubicacion_fisica,
                    estado_validacion, usuario_carga_id, notas
                ) VALUES (
                    :empresa_id, :tipo_id, :numero, :fecha, :fecha_venc,
                    :cliente_id, :proveedor_id,
                    :subtotal, :iva, :total, :saldo_inicial,
                    'ocr_imagen', :condicion, :medio_contado, :ubic,
                    'confirmado', :usuario_id, :notas
                )
                RETURNING id, numero_comprobante, fecha_emision,
                          monto_total, saldo_pendiente, estado_validacion, fecha_creacion
            """),
            {
                "empresa_id": empresa_id,
                "tipo_id": tipo_id,
                "numero": datos.numero_comprobante.strip(),
                "fecha": fecha_obj,
                "fecha_venc": fecha_venc_obj,
                "cliente_id": cliente_id,
                "proveedor_id": proveedor_id,
                "subtotal": monto_subtotal,
                "iva": monto_iva,
                "total": monto_total,
                "saldo_inicial": Decimal("0") if datos.condicion == "contado" else monto_total,
                "condicion": datos.condicion,
                "medio_contado": (datos.medio_pago_contado or "efectivo") if datos.condicion == "contado" else None,
                "ubic": (datos.ubicacion_fisica.strip() if datos.ubicacion_fisica else None),
                "usuario_id": usuario_id,
                "notas": notas,
            },
        )
        comprobante = dict(result.mappings().first())
    except IntegrityError as e:
        await db.rollback()
        msg = str(e.orig) if e.orig else str(e)
        if "numero_comprobante" in msg or "UniqueViolation" in type(e.orig).__name__:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Ya existe una {tipo_nombre.lower()} con número "
                    f"'{datos.numero_comprobante.strip()}'. Cada número de "
                    "comprobante debe ser único por emisor/tipo."
                ),
            )
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {msg[:200]}")

    # 5) Insertar items si hay (con cruce código ↔ descripción contra inventario)
    for item in datos.items:
        cantidad = Decimal(str(item.cantidad or 1))
        precio = Decimal(str(item.precio_unitario or 0))
        iva_pct = Decimal(str(item.porcentaje_iva if item.porcentaje_iva in (0, 5, 10) else 10))
        subtotal = (cantidad * precio).quantize(Decimal("0.01"))
        if iva_pct == Decimal("10"):
            iva_monto = (subtotal / Decimal("11")).quantize(Decimal("0.01"))
        elif iva_pct == Decimal("5"):
            iva_monto = (subtotal / Decimal("21")).quantize(Decimal("0.01"))
        else:
            iva_monto = Decimal("0.00")

        # Cruce con catálogo: si OCR leyó solo código o solo descripción, completamos
        inv_id, _codigo_final, desc_final = await _resolver_articulo(
            db, empresa_id, item.codigo, item.descripcion
        )

        await db.execute(
            text("""
                INSERT INTO detalle_comprobantes (
                    empresa_id, comprobante_id, inventario_id, descripcion,
                    cantidad, precio_unitario, porcentaje_iva,
                    subtotal, iva_monto
                ) VALUES (
                    :empresa_id, :comprobante_id, :inv_id, :desc,
                    :cantidad, :precio, :iva_pct,
                    :subtotal, :iva_monto
                )
            """),
            {
                "empresa_id": empresa_id,
                "comprobante_id": str(comprobante["id"]),
                "inv_id": inv_id,
                "desc": (desc_final or "Ítem").strip()[:300],
                "cantidad": cantidad,
                "precio": precio,
                "iva_pct": iva_pct,
                "subtotal": subtotal,
                "iva_monto": iva_monto,
            },
        )

    await db.commit()

    return {
        "ok": True,
        "mensaje": f"{tipo_nombre} N° {datos.numero_comprobante} guardada correctamente.",
        "comprobante_id": str(comprobante["id"]),
        "comprobante": {
            "numero_comprobante": comprobante["numero_comprobante"],
            "fecha_emision": str(comprobante["fecha_emision"]),
            "monto_total": str(comprobante["monto_total"]),
            "estado": comprobante["estado_validacion"],
        },
        "contraparte": {
            "tipo": "cliente" if datos.tipo == "venta" else "proveedor",
            "id": cliente_id or proveedor_id,
            "nombre": (datos.razon_social_cliente if datos.tipo == "venta" else datos.razon_social_emisor) or "",
            "ruc": (datos.ruc_cliente if datos.tipo == "venta" else datos.ruc_emisor) or "",
        },
    }


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTACIÓN MASIVA DESDE EXCEL
# Pensado para la carga inicial: una planilla con una factura por fila.
# Columnas esperadas (flexibles — case-insensitive, acepta sinónimos):
#   tipo            (venta / compra)         — default: venta
#   numero_comprobante                       — obligatorio
#   fecha_emision   (YYYY-MM-DD o dd/mm/yyyy)— obligatorio
#   ruc_cliente     (si tipo=venta)
#   razon_social_cliente
#   ruc_emisor      (si tipo=compra)
#   razon_social_emisor
#   condicion       (contado/credito)        — default: contado
#   monto_subtotal, monto_iva_5, monto_iva_10, monto_total
#   descripcion_item, cantidad, precio_unitario, porcentaje_iva, codigo (opcional)
# ═══════════════════════════════════════════════════════════════════════════════


# Sinónimos de columnas (todo en minúscula, sin tildes, sin espacios)
_ALIASES: dict[str, list[str]] = {
    "tipo": ["tipo", "tipo_factura", "tipodefactura"],
    "numero": ["numero", "numerocomprobante", "numero_factura", "nrofactura", "nro", "n"],
    "fecha": ["fecha", "fechaemision", "fecha_emision", "fechadelafactura"],
    "ruc_cliente": ["ruccliente", "ruc_cliente", "rucdelcliente"],
    "nombre_cliente": ["razonsocialcliente", "razon_social_cliente", "cliente", "nombrecliente"],
    "ruc_emisor": ["rucemisor", "ruc_emisor", "rucproveedor", "ruc_proveedor"],
    "nombre_emisor": ["razonsocialemisor", "razon_social_emisor", "emisor", "proveedor", "nombreemisor"],
    "condicion": ["condicion", "condiciondeventa"],
    "medio_pago": ["mediopago", "medio_pago", "mediopagocontado", "medio_pago_contado"],
    "fecha_vencimiento": ["fechavencimiento", "fecha_vencimiento", "vencimiento"],
    "ubicacion_fisica": ["ubicacionfisica", "ubicacion_fisica", "archivo_fisico", "archivador"],
    "subtotal": ["subtotal", "monto_subtotal", "montosubtotal"],
    "exentas": ["exentas", "exento", "monto_exentas", "iva0", "iva_0"],
    "iva_5": ["iva5", "iva_5", "monto_iva_5", "iva5pct"],
    "iva_10": ["iva10", "iva_10", "monto_iva_10", "iva10pct"],
    "total": ["total", "monto_total", "montototal", "importetotal"],
    "item_desc": ["descripcionitem", "descripcion", "articulo", "producto"],
    "item_cod": ["codigo", "codigoitem", "codarticulo"],
    "item_qty": ["cantidad", "cant", "qty"],
    "item_precio": ["preciounitario", "precio_unitario", "precio", "pu"],
    "item_iva": ["porcentajeiva", "porcentaje_iva", "ivaitem", "ivapct"],
}

PLANTILLA_HEADERS = [
    "tipo",
    "numero_comprobante",
    "fecha_emision",
    "ruc_cliente",
    "razon_social_cliente",
    "ruc_emisor",
    "razon_social_emisor",
    "condicion",
    "medio_pago",
    "fecha_vencimiento",
    "monto_subtotal",
    "monto_iva_5",
    "monto_iva_10",
    "monto_total",
    "descripcion_item",
    "cantidad",
    "precio_unitario",
    "porcentaje_iva",
    "codigo",
    "ubicacion_fisica",
]


def _normalizar_col(nombre: str) -> str:
    """Pasa 'RUC Cliente' → 'ruccliente' para matchear contra los aliases."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(nombre or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _map_headers(headers: list[str]) -> dict[str, int]:
    """Mapea cada alias al índice de columna correspondiente. Devuelve {clave: índice}."""
    normalizadas = [_normalizar_col(h) for h in headers]
    out: dict[str, int] = {}
    for clave, aliases in _ALIASES.items():
        for i, col in enumerate(normalizadas):
            if col in aliases:
                out[clave] = i
                break
    return out


@router.get("/plantilla-excel", summary="Descargar modelo Excel para carga masiva")
async def descargar_plantilla_excel(
    current_user: dict = Depends(get_current_user),
):
    """Genera una plantilla .xlsx vacia con el formato aceptado por la importacion."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ejemplos = wb.create_sheet("Ejemplos")
    ayuda = wb.create_sheet("Ayuda")

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(PLANTILLA_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ejemplos.append(PLANTILLA_HEADERS)
    ejemplos.append([
        "venta", "001-001-0000001", date.today(), "80000000-1", "Cliente Ejemplo SA",
        "", "", "credito", "transferencia", date.today(), 100000, 0, 10000, 110000,
        "Producto o servicio", 1, 110000, 10, "SKU-001", "Carpeta A",
    ])
    ejemplos.append([
        "compra", "001-001-0000123", date.today(), "", "",
        "80011111-2", "Proveedor Ejemplo SRL", "contado", "efectivo", "", 50000, 0, 5000, 55000,
        "Compra de insumos", 1, 55000, 10, "", "Caja comprobantes",
    ])

    widths = [14, 22, 16, 18, 28, 18, 28, 14, 18, 18, 16, 14, 14, 16, 28, 12, 16, 14, 14, 22]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        ejemplos.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PLANTILLA_HEADERS))}1"
    ejemplos.freeze_panes = "A2"
    ejemplos.auto_filter.ref = f"A1:{get_column_letter(len(PLANTILLA_HEADERS))}1"
    for cell in ejemplos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ejemplos.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = border

    tipo_dv = DataValidation(type="list", formula1='"venta,compra"', allow_blank=False)
    condicion_dv = DataValidation(type="list", formula1='"contado,credito"', allow_blank=True)
    medio_dv = DataValidation(type="list", formula1='"efectivo,transferencia,cheque,tarjeta,otro"', allow_blank=True)
    iva_dv = DataValidation(type="list", formula1='"0,5,10"', allow_blank=True)
    ws.add_data_validation(tipo_dv)
    ws.add_data_validation(condicion_dv)
    ws.add_data_validation(medio_dv)
    ws.add_data_validation(iva_dv)
    tipo_dv.add("A2:A500")
    condicion_dv.add("H2:H500")
    medio_dv.add("I2:I500")
    iva_dv.add("R2:R500")

    ayuda.append(["Columna", "Uso"])
    for row in [
        ("tipo", "venta o compra. En ventas se usa cliente; en compras se usa proveedor/emisor."),
        ("numero_comprobante", "Obligatorio. No debe repetirse."),
        ("fecha_emision", "Obligatorio. Formato recomendado: YYYY-MM-DD o fecha de Excel."),
        ("ruc_cliente / razon_social_cliente", "Completar para ventas. Si el cliente no existe, se crea."),
        ("ruc_emisor / razon_social_emisor", "Completar para compras. Si el proveedor no existe, se crea."),
        ("condicion", "contado o credito. Si se deja vacio se toma contado."),
        ("medio_pago", "Solo para contado: efectivo, transferencia, cheque, tarjeta u otro."),
        ("fecha_vencimiento", "Opcional para credito."),
        ("monto_subtotal / IVA / total", "El total es obligatorio. Si subtotal queda vacio, se calcula como total - IVA."),
        ("descripcion_item/cantidad/precio_unitario/porcentaje_iva/codigo", "Opcional. Si se carga, crea el detalle del comprobante."),
        ("ubicacion_fisica", "Opcional. Donde queda guardado el papel: carpeta, caja o archivador."),
    ]:
        ayuda.append(row)
    for cell in ayuda[1]:
        cell.fill = header_fill
        cell.font = header_font
    ayuda.column_dimensions["A"].width = 28
    ayuda.column_dimensions["B"].width = 90
    ayuda.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="modelo_carga_facturas.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _parsear_fecha(valor) -> Optional[date]:
    """Acepta date, datetime, string 'YYYY-MM-DD' o 'dd/mm/yyyy'."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, date) and not isinstance(valor, datetime):
        return valor
    if isinstance(valor, datetime):
        return valor.date()
    s = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parsear_monto(valor) -> Decimal:
    """Tolerante a strings con puntos/comas como separadores de miles."""
    if valor is None or valor == "":
        return Decimal("0")
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    s = str(valor).strip().replace(" ", "")
    # Si tiene coma decimal (ej. "1.190.000,50") → quitar puntos, cambiar coma por punto
    if "," in s and s.rfind(",") > s.rfind("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        # Formato americano o sin decimales (ej. "1,190,000" o "1190000")
        s = s.replace(",", "")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


@router.post("/importar-excel", summary="Importación masiva de facturas desde Excel")
async def importar_excel(
    archivo: UploadFile = File(..., description="Archivo .xlsx con una factura por fila"),
    tipo_default: Literal["venta", "compra"] = Query("venta"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_escritura),
):
    """
    Importa un lote de facturas desde una planilla Excel (una fila = una factura).

    - Detecta automáticamente las columnas por nombre (tolerante a mayúsculas/tildes/sinónimos).
    - Ignora filas vacías y filas de cabecera repetidas.
    - Salta duplicados (número ya existente) sin cortar el lote.
    - Autocrea cliente/proveedor si el RUC no existe.
    - Retorna un resumen: creadas, duplicadas, errores.
    """
    empresa_id = current_user["empresa_id"]
    usuario_id = current_user["sub"]

    # Validar MIME / extensión
    nombre = (archivo.filename or "").lower()
    if not nombre.endswith(".xlsx") and archivo.content_type not in MIME_EXCEL:
        raise HTTPException(status_code=422, detail="Formato no soportado. Subí un archivo .xlsx")

    contenido = await archivo.read()
    if len(contenido) / (1024 * 1024) > MAX_SIZE_MB:
        raise HTTPException(status_code=413, detail=f"Archivo muy grande (>{MAX_SIZE_MB}MB)")

    # Abrir con openpyxl
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo Excel: {e}")

    if ws is None:
        raise HTTPException(status_code=422, detail="La planilla no tiene hojas")

    rows_iter = ws.iter_rows(values_only=True)
    # Leer header de la primera fila no vacía
    headers: list = []
    for row in rows_iter:
        if row and any(c is not None and str(c).strip() for c in row):
            headers = [str(c) if c is not None else "" for c in row]
            break

    if not headers:
        raise HTTPException(status_code=422, detail="La planilla está vacía")

    mapa = _map_headers(headers)
    faltantes = [c for c in ("numero", "fecha", "total") if c not in mapa]
    if faltantes:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Faltan columnas obligatorias: {', '.join(faltantes)}. "
                f"Headers detectados: {headers}"
            ),
        )

    def cel(row, clave):
        idx = mapa.get(clave)
        return row[idx] if idx is not None and idx < len(row) else None

    creadas: list[dict] = []
    duplicadas: list[dict] = []
    errores: list[dict] = []

    for fila_idx, row in enumerate(rows_iter, start=2):
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue

        numero = str(cel(row, "numero") or "").strip()
        if not numero:
            continue  # fila sin número: probablemente fila en blanco con algún cálculo suelto

        fecha_obj = _parsear_fecha(cel(row, "fecha"))
        if not fecha_obj:
            errores.append({"fila": fila_idx, "numero": numero, "error": "Fecha inválida o vacía"})
            continue

        tipo_val = str(cel(row, "tipo") or tipo_default).strip().lower()
        if tipo_val not in ("venta", "compra"):
            tipo_val = tipo_default
        tipo_nombre = "Factura de Venta" if tipo_val == "venta" else "Factura de Compra"

        try:
            tipo_id = await _obtener_tipo_id(db, empresa_id, tipo_nombre)
            existe = (await db.execute(
                text("""
                    SELECT id
                    FROM comprobantes
                    WHERE empresa_id = :e AND numero_comprobante = :n
                    LIMIT 1
                """),
                {"e": empresa_id, "n": numero},
            )).mappings().first()
            if existe:
                duplicadas.append({"fila": fila_idx, "numero": numero})
                continue

            ruc_cli = _limpiar_ruc(str(cel(row, "ruc_cliente") or "") or None)
            nom_cli = str(cel(row, "nombre_cliente") or "").strip() or None
            ruc_emi = _limpiar_ruc(str(cel(row, "ruc_emisor") or "") or None)
            nom_emi = str(cel(row, "nombre_emisor") or "").strip() or None

            cliente_id = None
            proveedor_id = None
            if tipo_val == "venta":
                cliente_id = await _resolver_cliente(db, empresa_id, nom_cli, ruc_cli, usuario_id)
                if not cliente_id:
                    errores.append({"fila": fila_idx, "numero": numero, "error": "Falta RUC o razón social del cliente"})
                    continue
            else:
                proveedor_id = await _resolver_proveedor(db, empresa_id, nom_emi, ruc_emi, usuario_id)
                if not proveedor_id:
                    errores.append({"fila": fila_idx, "numero": numero, "error": "Falta RUC o razón social del proveedor"})
                    continue

            total = _parsear_monto(cel(row, "total"))
            subtotal = _parsear_monto(cel(row, "subtotal"))
            exentas = _parsear_monto(cel(row, "exentas"))
            iva_5 = _parsear_monto(cel(row, "iva_5"))
            iva_10 = _parsear_monto(cel(row, "iva_10"))
            iva = iva_5 + iva_10
            condicion = str(cel(row, "condicion") or "contado").strip().lower()
            if condicion not in ("contado", "credito"):
                condicion = "contado"
            medio_pago = str(cel(row, "medio_pago") or "efectivo").strip().lower()
            if medio_pago not in ("efectivo", "transferencia", "cheque", "tarjeta", "otro"):
                medio_pago = "efectivo"
            fecha_venc = _parsear_fecha(cel(row, "fecha_vencimiento")) if condicion == "credito" else None
            ubicacion = str(cel(row, "ubicacion_fisica") or "").strip() or None
            # Si viene el subtotal, lo respetamos. Si no, lo deducimos: total - IVA.
            # Los exentos ya están incluidos en el subtotal (subtotal = base_5 + base_10 + exentas).
            if subtotal == 0 and total > 0:
                subtotal = total - iva
            # Si sólo nos dieron exentas sueltas (sin subtotal explícito), mínimo garantizamos
            # que el subtotal no sea menor que los exentos.
            if exentas > subtotal:
                subtotal = exentas

            try:
                result = await db.execute(
                    text("""
                        INSERT INTO comprobantes (
                            empresa_id, tipo_id, numero_comprobante, fecha_emision,
                            fecha_vencimiento, cliente_id, proveedor_id,
                            monto_subtotal, monto_iva, monto_total, saldo_pendiente,
                            metodo_carga, condicion, medio_pago_contado, ubicacion_fisica,
                            estado_validacion, usuario_carga_id, notas
                        ) VALUES (
                            :e, :t, :n, :f, :fecha_venc, :cli, :prov,
                            :sub, :iva, :tot, :saldo,
                            'manual', :condicion, :medio_pago, :ubicacion,
                            'confirmado', :u,
                            'Importado desde Excel'
                        )
                        RETURNING id
                    """),
                    {
                        "e": empresa_id, "t": tipo_id, "n": numero, "f": fecha_obj,
                        "fecha_venc": fecha_venc,
                        "cli": cliente_id, "prov": proveedor_id,
                        "sub": subtotal, "iva": iva, "tot": total,
                        "saldo": Decimal("0") if condicion == "contado" else total,
                        "condicion": condicion,
                        "medio_pago": medio_pago if condicion == "contado" else None,
                        "ubicacion": ubicacion,
                        "u": usuario_id,
                    },
                )
                comp_id = str(result.mappings().first()["id"])
            except IntegrityError as e:
                await db.rollback()
                msg = f"{type(e.orig).__name__ if e.orig else ''} {str(e.orig) if e.orig else str(e)}"
                if "UniqueViolation" in msg or "numero_comprobante" in msg or "duplicada" in msg.lower():
                    duplicadas.append({"fila": fila_idx, "numero": numero})
                    continue
                errores.append({"fila": fila_idx, "numero": numero, "error": f"DB: {msg[:120]}"})
                continue

            # Item único por fila (si existe)
            desc_item = str(cel(row, "item_desc") or "").strip()
            cod_item = str(cel(row, "item_cod") or "").strip() or None
            if desc_item or cod_item:
                cantidad = Decimal(str(cel(row, "item_qty") or 1))
                precio = _parsear_monto(cel(row, "item_precio"))
                iva_pct_raw = cel(row, "item_iva")
                try:
                    iva_pct = int(float(iva_pct_raw)) if iva_pct_raw is not None else 10
                except Exception:
                    iva_pct = 10
                if iva_pct not in (0, 5, 10):
                    iva_pct = 10
                if precio == 0 and total > 0 and cantidad > 0:
                    precio = total / cantidad

                inv_id, _, desc_final = await _resolver_articulo(db, empresa_id, cod_item, desc_item or None)
                sub_item = (cantidad * precio).quantize(Decimal("0.01"))
                if iva_pct == 10:
                    iva_monto = (sub_item / Decimal("11")).quantize(Decimal("0.01"))
                elif iva_pct == 5:
                    iva_monto = (sub_item / Decimal("21")).quantize(Decimal("0.01"))
                else:
                    iva_monto = Decimal("0.00")

                await db.execute(
                    text("""
                        INSERT INTO detalle_comprobantes (
                            empresa_id, comprobante_id, inventario_id, descripcion,
                            cantidad, precio_unitario, porcentaje_iva,
                            subtotal, iva_monto
                        ) VALUES (
                            :e, :c, :inv, :d, :q, :p, :iv, :sub, :ivm
                        )
                    """),
                    {
                        "e": empresa_id, "c": comp_id, "inv": inv_id,
                        "d": (desc_final or "Ítem").strip()[:300],
                        "q": cantidad, "p": precio, "iv": iva_pct,
                        "sub": sub_item, "ivm": iva_monto,
                    },
                )

            creadas.append({"fila": fila_idx, "numero": numero, "id": comp_id})

        except Exception as e:
            await db.rollback()
            errores.append({"fila": fila_idx, "numero": numero, "error": str(e)[:200]})
            continue

    await db.commit()

    return {
        "resumen": {
            "creadas": len(creadas),
            "duplicadas": len(duplicadas),
            "errores": len(errores),
            "total_filas_procesadas": len(creadas) + len(duplicadas) + len(errores),
        },
        "creadas": creadas,
        "duplicadas": duplicadas,
        "errores": errores,
    }
