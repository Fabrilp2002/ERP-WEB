"""
Servicio de Exportación Excel — openpyxl.
Genera archivos .xlsx con formato profesional para reportes del ERP.
"""
import io
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


# ── Estilos ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
CURRENCY_FORMAT = '#,##0'  # Guaraníes sin decimales
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _auto_width(ws):
    """Ajusta automáticamente el ancho de columnas.
    Robusto frente a MergedCell (filas de título) que no exponen column_letter."""
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    max_lens: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None:
                continue
            col_idx = cell.column  # 1-based int
            val = str(cell.value)
            if len(val) > max_lens.get(col_idx, 0):
                max_lens[col_idx] = len(val)
    for col_idx, length in max_lens.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 3, 45)


def _add_header_row(ws, headers: list[str], row: int = 1):
    """Agrega fila de encabezado con estilo."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _format_guarani(value) -> int:
    """Convierte Decimal/str/float a entero guaraní."""
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return int(value)
    try:
        return int(Decimal(str(value)))
    except Exception:
        return 0


# ── Exportar Comprobantes ─────────────────────────────────────────────────────

def generar_excel_comprobantes(comprobantes: list[dict], empresa_nombre: str = "ERP") -> bytes:
    """Genera Excel con listado de comprobantes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    # Título
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"Comprobantes — {empresa_nombre}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")

    # Encabezados (ahora con emisor/imputado = contraparte)
    headers = [
        "N° Comprobante", "Fecha Emisión", "Vencimiento", "Tipo",
        "Contraparte", "RUC", "Condición",
        "Subtotal", "IVA", "Total", "Saldo Pendiente",
        "Estado", "Ubicación física",
    ]
    _add_header_row(ws, headers, row=4)

    # Datos
    for i, c in enumerate(comprobantes, start=5):
        ws.cell(row=i, column=1, value=c.get("numero_comprobante", ""))
        ws.cell(row=i, column=2, value=str(c.get("fecha_emision", "") or ""))
        ws.cell(row=i, column=3, value=str(c.get("fecha_vencimiento", "") or ""))
        ws.cell(row=i, column=4, value=c.get("tipo", ""))
        ws.cell(row=i, column=5, value=c.get("contraparte", ""))
        ws.cell(row=i, column=6, value=c.get("contraparte_ruc", ""))
        ws.cell(row=i, column=7, value=c.get("condicion", ""))

        for col_idx, field in [(8, "monto_subtotal"), (9, "monto_iva"),
                                (10, "monto_total"), (11, "saldo_pendiente")]:
            cell = ws.cell(row=i, column=col_idx, value=_format_guarani(c.get(field, 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = Alignment(horizontal="right")

        ws.cell(row=i, column=12, value=c.get("estado_validacion", ""))
        ws.cell(row=i, column=13, value=c.get("ubicacion_fisica", "") or "")

        for col_idx in range(1, 14):
            ws.cell(row=i, column=col_idx).border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Exportar Cuentas Corrientes ───────────────────────────────────────────────

def generar_excel_cuentas_corrientes(
    clientes: list[dict],
    proveedores: list[dict],
    empresa_nombre: str = "ERP",
) -> bytes:
    """Genera Excel con saldos de clientes y proveedores en hojas separadas."""
    wb = Workbook()

    # Hoja Clientes
    ws_cli = wb.active
    ws_cli.title = "Clientes"
    ws_cli.merge_cells("A1:E1")
    ws_cli["A1"].value = f"Cuentas Corrientes — Clientes — {empresa_nombre}"
    ws_cli["A1"].font = TITLE_FONT

    _add_header_row(ws_cli, ["Cliente", "Total Facturado", "Total Cobrado", "Saldo Pendiente"], row=3)

    for i, c in enumerate(clientes, start=4):
        ws_cli.cell(row=i, column=1, value=c.get("cliente", ""))
        for col_idx, field in [(2, "total_facturado"), (3, "total_cobrado"), (4, "saldo_pendiente")]:
            cell = ws_cli.cell(row=i, column=col_idx, value=_format_guarani(c.get(field, 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = Alignment(horizontal="right")
        for col_idx in range(1, 5):
            ws_cli.cell(row=i, column=col_idx).border = THIN_BORDER

    _auto_width(ws_cli)

    # Hoja Proveedores
    ws_prov = wb.create_sheet("Proveedores")
    ws_prov.merge_cells("A1:E1")
    ws_prov["A1"].value = f"Cuentas Corrientes — Proveedores — {empresa_nombre}"
    ws_prov["A1"].font = TITLE_FONT

    _add_header_row(ws_prov, ["Proveedor", "Total Facturado", "Total Pagado", "Saldo Pendiente"], row=3)

    for i, p in enumerate(proveedores, start=4):
        ws_prov.cell(row=i, column=1, value=p.get("proveedor", ""))
        for col_idx, field in [(2, "total_facturado"), (3, "total_pagado"), (4, "saldo_pendiente")]:
            cell = ws_prov.cell(row=i, column=col_idx, value=_format_guarani(p.get(field, 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = Alignment(horizontal="right")
        for col_idx in range(1, 5):
            ws_prov.cell(row=i, column=col_idx).border = THIN_BORDER

    _auto_width(ws_prov)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Exportar Movimientos (cobros + pagos) ─────────────────────────────────────

def _hoja_movimientos(ws, rows: list[dict], titulo: str, empresa_nombre: str, color_total: str):
    """Pinta una hoja de cobros o pagos con todos los detalles y un total al final."""
    ws.merge_cells("A1:J1")
    ws["A1"].value = f"{titulo} — {empresa_nombre}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")

    headers = [
        "Fecha", "N° Recibo", "Contraparte", "RUC",
        "N° Comprobante", "Medio de Pago", "Monto", "Usuario", "Notas",
    ]
    _add_header_row(ws, headers, row=4)

    total = Decimal(0)
    for i, r in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=str(r.get("fecha_pago", "") or ""))
        ws.cell(row=i, column=2, value=r.get("numero_recibo", "") or "")
        ws.cell(row=i, column=3, value=r.get("contraparte", "") or "—")
        ws.cell(row=i, column=4, value=r.get("contraparte_ruc", "") or "")
        ws.cell(row=i, column=5, value=r.get("numero_comprobante", "") or "")
        ws.cell(row=i, column=6, value=r.get("medio_pago", "") or "")
        monto_cell = ws.cell(row=i, column=7, value=_format_guarani(r.get("monto_pagado", 0)))
        monto_cell.number_format = CURRENCY_FORMAT
        monto_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=8, value=r.get("usuario", "") or "")
        ws.cell(row=i, column=9, value=r.get("notas", "") or "")
        for col_idx in range(1, 10):
            ws.cell(row=i, column=col_idx).border = THIN_BORDER
        try:
            total += Decimal(str(r.get("monto_pagado", 0) or 0))
        except Exception:
            pass

    # Fila de total
    total_row = 5 + len(rows) + 1
    ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
    tot_cell = ws.cell(row=total_row, column=7, value=_format_guarani(total))
    tot_cell.number_format = CURRENCY_FORMAT
    tot_cell.font = Font(bold=True)
    tot_cell.fill = PatternFill(start_color=color_total, end_color=color_total, fill_type="solid")
    tot_cell.alignment = Alignment(horizontal="right")

    _auto_width(ws)


def generar_excel_movimientos(
    cobros: list[dict],
    pagos: list[dict],
    empresa_nombre: str = "ERP",
) -> bytes:
    """
    Excel con 3 hojas:
      1. Cobros (ingresos de clientes)
      2. Pagos (egresos a proveedores)
      3. Resumen (totales + balance)
    Cada movimiento incluye al emisor/imputado (contraparte).
    """
    wb = Workbook()

    # Hoja 1: Cobros
    ws_cob = wb.active
    ws_cob.title = "Cobros"
    _hoja_movimientos(ws_cob, cobros, "Cobros (Ingresos)", empresa_nombre, "C6EFCE")

    # Hoja 2: Pagos
    ws_pag = wb.create_sheet("Pagos")
    _hoja_movimientos(ws_pag, pagos, "Pagos (Egresos)", empresa_nombre, "FFC7CE")

    # Hoja 3: Resumen
    ws_res = wb.create_sheet("Resumen")
    ws_res.merge_cells("A1:C1")
    ws_res["A1"].value = f"Resumen de Movimientos — {empresa_nombre}"
    ws_res["A1"].font = TITLE_FONT
    ws_res["A1"].alignment = Alignment(horizontal="center")

    total_cobros = sum((Decimal(str(r.get("monto_pagado", 0) or 0)) for r in cobros), Decimal(0))
    total_pagos = sum((Decimal(str(r.get("monto_pagado", 0) or 0)) for r in pagos), Decimal(0))
    balance = total_cobros - total_pagos

    _add_header_row(ws_res, ["Concepto", "Cantidad", "Total (₲)"], row=3)
    datos = [
        ("Cobros (ingresos)", len(cobros), total_cobros),
        ("Pagos (egresos)", len(pagos), total_pagos),
        ("BALANCE", len(cobros) + len(pagos), balance),
    ]
    for i, (concepto, cant, monto) in enumerate(datos, start=4):
        ws_res.cell(row=i, column=1, value=concepto)
        ws_res.cell(row=i, column=2, value=cant)
        cell = ws_res.cell(row=i, column=3, value=_format_guarani(monto))
        cell.number_format = CURRENCY_FORMAT
        cell.alignment = Alignment(horizontal="right")
        if concepto == "BALANCE":
            color = "C6EFCE" if balance >= 0 else "FFC7CE"
            for col in range(1, 4):
                c = ws_res.cell(row=i, column=col)
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 4):
            ws_res.cell(row=i, column=col).border = THIN_BORDER

    _auto_width(ws_res)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Exportar Inventario ───────────────────────────────────────────────────────

def generar_excel_inventario(items: list[dict], empresa_nombre: str = "ERP") -> bytes:
    """Genera Excel con listado de inventario + alertas de stock."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.merge_cells("A1:G1")
    ws["A1"].value = f"Inventario — {empresa_nombre}"
    ws["A1"].font = TITLE_FONT

    headers = ["Código", "Descripción", "Unidad", "Stock Actual", "Punto Reorden", "Costo Unit.", "Estado"]
    _add_header_row(ws, headers, row=3)

    ALERT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    CRITICAL_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for i, item in enumerate(items, start=4):
        ws.cell(row=i, column=1, value=item.get("codigo", ""))
        ws.cell(row=i, column=2, value=item.get("descripcion", ""))
        ws.cell(row=i, column=3, value=item.get("unidad_medida", ""))

        cant = float(item.get("cantidad_actual", 0))
        reorden = float(item.get("punto_reorden", 0))

        ws.cell(row=i, column=4, value=cant)
        ws.cell(row=i, column=5, value=reorden)

        costo_cell = ws.cell(row=i, column=6, value=_format_guarani(item.get("costo_unitario", 0)))
        costo_cell.number_format = CURRENCY_FORMAT

        # Estado de stock
        if reorden > 0 and cant <= reorden:
            if cant == 0:
                estado = "SIN STOCK"
                fill = CRITICAL_FILL
            else:
                estado = "BAJO STOCK"
                fill = ALERT_FILL
        else:
            estado = "Normal"
            fill = None

        estado_cell = ws.cell(row=i, column=7, value=estado)
        if fill:
            for col_idx in range(1, 8):
                ws.cell(row=i, column=col_idx).fill = fill

        for col_idx in range(1, 8):
            ws.cell(row=i, column=col_idx).border = THIN_BORDER

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_excel_iva(tipo: str, filas: list[dict], totales: dict, empresa_nombre: str = "ERP", periodo: str = "") -> bytes:
    """Genera Libro IVA ventas/compras en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IVA Ventas" if tipo == "ventas" else "IVA Compras"

    headers = (
        ["RUC Comprador", "Razon Social", "Nro. Comprobante", "Fecha", "Tipo",
         "Base 10%", "IVA 10%", "Base 5%", "IVA 5%", "Exentas", "Total"]
        if tipo == "ventas"
        else ["RUC Proveedor", "Razon Social", "Nro. Comprobante", "Fecha", "Tipo",
              "Base 10%", "Credito 10%", "Base 5%", "Credito 5%", "Exentas", "Total"]
    )
    fields = (
        ["ruc_comprador", "razon_social", "numero_comprobante", "fecha_emision", "tipo_comprobante",
         "base_gravada_10", "iva_10", "base_gravada_5", "iva_5", "exentas", "total"]
        if tipo == "ventas"
        else ["ruc_proveedor", "razon_social", "numero_comprobante", "fecha_emision", "tipo_comprobante",
              "base_gravada_10", "credito_fiscal_10", "base_gravada_5", "credito_fiscal_5", "exentas", "total"]
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value=f"Libro IVA {tipo.title()} - {empresa_nombre}")
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1, value=f"Periodo: {periodo or 'Todo'}")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")

    _add_header_row(ws, headers, row=4)
    money_fields = set(fields[5:])
    for row_idx, fila in enumerate(filas, start=5):
        for col_idx, field in enumerate(fields, start=1):
            value = fila.get(field)
            cell = ws.cell(row=row_idx, column=col_idx)
            if field in money_fields:
                cell.value = _format_guarani(value)
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = str(value or "")
            cell.border = THIN_BORDER

    total_row = 5 + len(filas) + 1
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
    for col_idx, field in enumerate(fields, start=1):
        if field in money_fields:
            cell = ws.cell(row=total_row, column=col_idx, value=_format_guarani(totales.get(field, 0)))
            cell.number_format = CURRENCY_FORMAT
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    _auto_width(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
